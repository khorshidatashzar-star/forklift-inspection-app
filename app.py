# BUILD: SIMPLE-V30-SUPABASE-PHOTO-STORAGE-20260820
# BUILD: SIMPLE-V29-EXCEL-EXPORT-FIX-20260820\n# BUILD: SIMPLE-V28-DRIVER-SPEED-POOL-20260820
# BUILD: SIMPLE-V27-PERFORMANCE-OPTIMIZED-20260820
# BUILD: SIMPLE-V26-ROLE-LOCATION-UI-FIX-20260820
# BUILD: SIMPLE-V25-ROLE-ACCESS-ENFORCED-20260820
# BUILD: SIMPLE-V24-FULL-MACHINE-HISTORY-20260820
# BUILD: SIMPLE-V23-STAFF-FIRST-LOGIN-FIX-20260820
# BUILD: SIMPLE-V22-SUPABASE-EMPTY-QUERY-FIX-20260820\n# BUILD: SIMPLE-V21-SUPABASE-DATABASE-20260820
# BUILD: SIMPLE-V20-CLOUD-STABLE-ALL-EXCEL-20260819
# BUILD: SIMPLE-V19-IT-TEST-PORTAL-20260819
# BUILD: SIMPLE-V18-IT-ADMIN-CLEANUP-20260819
# BUILD: SIMPLE-V17-CLOUD-EXCEL-FIX-20260819
# BUILD: SIMPLE-V16-CLOUD-SAFE
# BUILD: SIMPLE-V15-MANAGER-COMPLIANCE-TABS-20260819
# BUILD: SIMPLE-V14-MAINTENANCE-TABS-20260819
# BUILD: SIMPLE-V13-SHIFT-REPORTS-MACHINE-HISTORY-20260819
# BUILD: SIMPLE-V12-SHIFT-FIX-CLEAN-20260819
# BUILD: SIMPLE-V11-SHIFT-DB-FIXED-20260819
# BUILD: SIMPLE-V11-SHIFTS-MACHINE-HISTORY-20260819
# BUILD: SIMPLE-V10-HUMAN-LABELS-TIMELINE-20260819
# BUILD: SIMPLE-V9-SUBMIT-MESSAGE-UI-20260819
# BUILD: SIMPLE-V8-DOWNLOAD-TODAY-STATUS-20260818
# BUILD: SIMPLE-V7-DAILY-TASKS-CASE-HISTORY-20260818
# BUILD: SIMPLE-V5-COMPLIANCE-TABS-20260818
# BUILD: SIMPLE-V5-CLEAN-REPORTS-20260818
# BUILD: SIMPLE-V5-RESET-AND-DETAILED-VALIDATION-20260818
# BUILD: SIMPLE-V5-FIXED-COLUMNS-20260818
# BUILD: SIMPLE-V5-20260818
# BUILD: V4-AUTO-CLOSE-REPAIR-20260818
# BUILD: V4-TEMP-ROLE-PREVIEW-20260817
# BUILD: V4-STAFF-ROW-ACTIONS-20260817

import streamlit as st
import sqlite3
import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2.pool import ThreadedConnectionPool
import json
import io
import re
import hashlib
import secrets
import smtplib
from pathlib import Path
from datetime import datetime, date, timedelta
from email.message import EmailMessage
import requests
import pandas as pd

BASE = Path(__file__).parent
DB = BASE / "forklift_role_portal_v4.db"
EQ_FILE = BASE / "equipment_master.csv"
UPLOAD = BASE / "uploads"
UPLOAD.mkdir(exist_ok=True)

# STAFF ROW ACTIONS ENABLED
st.set_page_config(
    page_title="18W Forklift Safety",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.block-container{max-width:1280px;padding-top:1rem;padding-bottom:4rem}
.hero{padding:18px 22px;border-radius:18px;background:linear-gradient(135deg,#f3f7fb,#fff);border:1px solid #dfe5ec;margin-bottom:18px}
.hero h2{margin:0 0 5px 0}
.muted{color:#667085;font-size:.9rem}
.section-title{font-weight:800;font-size:1.08rem;margin:1rem 0 .4rem}
.issue-card{border-left:7px solid #d92d20;background:#fff6f5;padding:12px 14px;border-radius:12px;margin:8px 0}
.ok-card{border-left:7px solid #12b76a;background:#f2fff6;padding:12px 14px;border-radius:12px;margin:8px 0}
.purple-card{border-left:7px solid #7f56d9;background:#faf7ff;padding:12px 14px;border-radius:12px;margin:8px 0}
.orange-card{border-left:7px solid #f79009;background:#fffaeb;padding:12px 14px;border-radius:12px;margin:8px 0}
.machine-selected{padding:7px 10px;border-radius:9px;background:#f4f7fb;border:1px solid #dbe4ef;font-size:.90rem;color:#344054;margin:.35rem 0 .65rem 0}
.stButton>button{border-radius:10px;font-weight:750;min-height:2.55rem}
[data-testid="stMetric"]{border:1px solid #e5e7eb;border-radius:12px;padding:12px;background:white}
</style>
""", unsafe_allow_html=True)

# Approved inspection grouping — keep this order unchanged.
GROUPS = [
    ("🔋 Battery & Electrical", [
        "Battery", "Vent Caps", "Connector Covers", "Cables", "Battery Connectors"
    ]),
    ("🏗️ Front / Mast / Forks", [
        "Overhead Guard", "Front Tire (Left)", "Tilt Cylinder (Left)", "Carriage (Left)",
        "Fork Locking Pin (Left)", "Fork (Left) / Attachment (if applicable)", "Mast",
        "Lift Cylinder", "Lift Chains", "Fork (Right) / Attachment (if applicable)",
        "Fork Locking Pin (Right)", "Carriage (Right)", "Tilt Cylinder (Right)", "Front Tire (Right)"
    ]),
    ("💺 Operator Area / Rear", [
        "Hydraulic Oil", "Data Plate", "Seat & Seat Belt", "Operator Manual",
        "Rear Tire (Right)", "Rear Tire (Left)"
    ]),
    ("🎛️ Controls & Brakes", [
        "Check Service & Parking Brake", "Lifting Control", "Tilt Control", "Gauges"
    ]),
    ("🚜 Drive Test", [
        "Forward Driving - Accelerator", "Forward Driving - Steering", "Forward Driving - Braking",
        "Reverse Driving - Accelerator", "Reverse Driving - Steering", "Reverse Driving - Braking",
        "Reverse Driving - Backup Alarm"
    ]),
    ("⚠️ Safety / General Operation", [
        "Listen for Unusual Noise", "Lights", "Horn", "Oil Spots on Floor"
    ]),
]

ISSUE_TYPES = {
    "Battery": ["Damaged", "Loose", "Corrosion", "Leak", "Low charge", "Other"],
    "Vent Caps": ["Missing", "Loose", "Cracked / damaged", "Blocked / dirty", "Leak / residue", "Corrosion", "Other"],
    "Connector Covers": ["Missing", "Loose", "Cracked / damaged", "Not secured", "Other"],
    "Cables": ["Loose", "Damaged", "Exposed wire", "Corrosion", "Poor connection", "Other"],
    "Battery Connectors": ["Loose", "Damaged", "Corrosion", "Poor connection", "Overheating / burn mark", "Other"],
    "Overhead Guard": ["Bent", "Cracked / damaged", "Loose", "Missing hardware", "Other"],
    "Carriage (Left)": ["Cracked / damaged", "Bent / deformed", "Loose", "Excessive movement", "Misaligned", "Not moving smoothly", "Other"],
    "Carriage (Right)": ["Cracked / damaged", "Bent / deformed", "Loose", "Excessive movement", "Misaligned", "Not moving smoothly", "Other"],
    "Lift Chains": ["Loose", "Damaged", "Excessive wear", "Rust / corrosion", "Uneven tension", "Other"],
    "Hydraulic Oil": ["Low level", "Leak", "Contaminated", "Other"],
    "Data Plate": ["Missing", "Damaged", "Unreadable", "Incorrect information", "Other"],
    "Seat & Seat Belt": ["Seat damaged", "Seat loose", "Seat belt missing", "Seat belt torn", "Seat belt not locking", "Other"],
    "Operator Manual": ["Missing", "Damaged / unreadable", "Wrong manual", "Other"],
    "Check Service & Parking Brake": ["Weak", "Not responding", "Parking brake not holding", "Unusual noise", "Other"],
    "Lifting Control": ["Not responding", "Slow response", "Sticking", "Jerky movement", "Other"],
    "Tilt Control": ["Not responding", "Slow response", "Sticking", "Jerky movement", "Other"],
    "Gauges": ["Not working", "Warning light", "Incorrect reading", "Damaged", "Other"],
    "Reverse Driving - Backup Alarm": ["Not working", "Low volume", "Intermittent", "Other"],
    "Listen for Unusual Noise": ["Grinding", "Squeaking", "Rattling", "Clicking", "Other"],
    "Lights": ["Not working", "Dim", "Broken", "Intermittent", "Other"],
    "Horn": ["Not working", "Low volume", "Intermittent", "Other"],
    "Oil Spots on Floor": ["Small leak", "Active leak", "Unknown fluid", "Other"],
}

DELAY_REASONS = [
    "Waiting for Part",
    "Vendor Required",
    "Machine Access Delayed",
    "Further Diagnosis Needed",
    "Other",
]

def generic_issue_types(component):
    if "Tire" in component:
        return ["Worn", "Cut", "Chunk missing", "Flat / low", "Damaged", "Other"]
    if "Cylinder" in component:
        return ["Leak", "Damaged", "Loose", "Not operating smoothly", "Other"]
    if "Fork" in component:
        return ["Bent", "Cracked", "Excessive wear", "Loose / not locking", "Damaged", "Other"]
    if "Steering" in component:
        return ["Loose", "Hard to steer", "Pulling", "Excessive play", "Other"]
    if "Accelerator" in component:
        return ["No response", "Delayed response", "Sticking", "Jerky operation", "Other"]
    if "Braking" in component:
        return ["Weak", "Delayed", "Pulling", "Unusual noise", "Other"]
    return ["Damaged", "Loose", "Not working correctly", "Excessive wear", "Missing", "Other"]

def _pg_sql(sql):
    # Existing app SQL uses SQLite-style ? placeholders.
    return sql.replace("?", "%s")


class PgCursor:
    def __init__(self, cursor):
        self.cursor = cursor

    def execute(self, sql, params=()):
        self.cursor.execute(_pg_sql(sql), params)
        return self

    def fetchone(self):
        return self.cursor.fetchone()

    def fetchall(self):
        return self.cursor.fetchall()

    def close(self):
        self.cursor.close()


@st.cache_resource
def pg_pool():
    if "DATABASE_URL" not in st.secrets:
        raise RuntimeError("DATABASE_URL is missing from Streamlit Secrets.")

    return ThreadedConnectionPool(
        minconn=1,
        maxconn=12,
        dsn=st.secrets["DATABASE_URL"],
        cursor_factory=RealDictCursor,
        connect_timeout=10,
        keepalives=1,
        keepalives_idle=30,
        keepalives_interval=10,
        keepalives_count=3,
    )


class PgConnection:
    def __init__(self):
        self.pool = pg_pool()
        self.raw = self.pool.getconn()

        # A pooled connection may have been left in an aborted transaction if
        # a previous rerun stopped mid-operation. Reset it before reuse.
        try:
            self.raw.rollback()
        except Exception:
            pass

    def execute(self, sql, params=()):
        cur = self.raw.cursor()
        cur.execute(_pg_sql(sql), params)
        return PgCursor(cur)

    def commit(self):
        self.raw.commit()

    def rollback(self):
        self.raw.rollback()

    def close(self):
        if self.raw is not None:
            self.pool.putconn(self.raw)
            self.raw = None


def conn():
    return PgConnection()

def now_dt():
    return datetime.now()

def now():
    return now_dt().strftime("%Y-%m-%d %H:%M:%S")

def pw_hash(password, salt=None):
    salt = salt or secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode(),
        salt.encode(),
        220000
    ).hex()
    return salt, h

def strong_password(p):
    return (
        len(p) >= 12
        and re.search(r"[A-Z]", p)
        and re.search(r"[a-z]", p)
        and re.search(r"\d", p)
        and re.search(r"[^A-Za-z0-9]", p)
    )

def init_db():
    """Verify the persistent Supabase PostgreSQL database is reachable."""
    c = conn()
    try:
        c.execute("SELECT 1").fetchone()
    finally:
        c.close()

def audit(case, role, name, event, details=""):
    c = conn()
    c.execute("""
    INSERT INTO audit(case_number,event_time,actor_role,actor_name,event_type,details)
    VALUES(?,?,?,?,?,?)
    """, (case or "", now(), role, name, event, details))
    c.commit()
    c.close()

def query(sql, params=()):
    c = conn()
    try:
        cur = c.raw.cursor()
        cur.execute(_pg_sql(sql), params)

        if cur.description is None:
            return pd.DataFrame()

        # Preserve column names even when the query returns zero rows.
        columns = [desc[0] for desc in cur.description]
        rows = cur.fetchall()

        if not rows:
            return pd.DataFrame(columns=columns)

        return pd.DataFrame(rows, columns=columns)
    finally:
        c.close()


@st.cache_data(ttl=60, show_spinner=False)
def equipment():
    d = query("""
        SELECT serial_number,unit_number,location,brand,description,
               model,status,building,out_of_service
        FROM equipment
        ORDER BY location,unit_number,serial_number
    """)

    cols = [
        "serial_number","unit_number","location","brand","description",
        "model","status","building","out_of_service"
    ]

    if d.empty:
        return pd.DataFrame(columns=cols)

    for col in cols:
        if col not in d.columns:
            d[col] = "No" if col == "out_of_service" else ""

    return d[cols].fillna("")


def save_equipment(d):
    """Persist equipment changes to Supabase instead of a local CSV."""
    cols = [
        "serial_number","unit_number","location","brand","description",
        "model","status","building","out_of_service"
    ]

    data = d.copy()
    for col in cols:
        if col not in data.columns:
            data[col] = ""

    c = conn()
    try:
        for _, r in data[cols].iterrows():
            serial = str(r["serial_number"] or "").strip()
            if not serial:
                continue

            c.execute("""
                INSERT INTO equipment(
                    serial_number,unit_number,location,brand,description,
                    model,status,building,out_of_service,updated_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,NOW())
                ON CONFLICT(serial_number)
                DO UPDATE SET
                    unit_number=EXCLUDED.unit_number,
                    location=EXCLUDED.location,
                    brand=EXCLUDED.brand,
                    description=EXCLUDED.description,
                    model=EXCLUDED.model,
                    status=EXCLUDED.status,
                    building=EXCLUDED.building,
                    out_of_service=EXCLUDED.out_of_service,
                    updated_at=NOW()
            """, (
                serial,
                str(r["unit_number"] or ""),
                str(r["location"] or ""),
                str(r["brand"] or ""),
                str(r["description"] or ""),
                str(r["model"] or ""),
                str(r["status"] or "Active"),
                str(r["building"] or ""),
                str(r["out_of_service"] or "No"),
            ))
        c.commit()
        equipment.clear()
    except Exception:
        c.rollback()
        raise
    finally:
        c.close()


def daily_case_number():
    prefix = f"FI-{date.today():%Y%m%d}-"
    c = conn()
    try:
        r = c.execute(
            """SELECT case_number
               FROM inspections
               WHERE case_number LIKE ?
               ORDER BY case_number DESC
               LIMIT 1""",
            (prefix + "%",)
        ).fetchone()
    finally:
        c.close()

    seq = 0
    if r:
        m = re.search(r"-(\d{3})$", r["case_number"] or "")
        if m:
            seq = int(m.group(1))

    return f"{prefix}{seq+1:03d}"

def _storage_config():
    required = ["SUPABASE_URL", "SUPABASE_SECRET_KEY"]
    missing = [k for k in required if not st.secrets.get(k)]
    if missing:
        raise RuntimeError(
            "Missing Streamlit Secret(s): " + ", ".join(missing)
        )

    return (
        str(st.secrets["SUPABASE_URL"]).rstrip("/"),
        str(st.secrets["SUPABASE_SECRET_KEY"]),
    )


def save_photo(upload, name):
    """Upload an inspection/repair photo to the private Supabase bucket."""
    if not upload:
        return ""

    base_url, secret_key = _storage_config()

    ext = Path(upload.name).suffix.lower() or ".jpg"
    mime = getattr(upload, "type", None) or {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
    }.get(ext, "application/octet-stream")

    # Keep objects organized by year/month and case/file name.
    object_path = f"{date.today():%Y/%m}/{name}{ext}"

    from urllib.parse import quote
    encoded_path = quote(object_path, safe="/")
    endpoint = (
        f"{base_url}/storage/v1/object/forklift-photos/{encoded_path}"
    )

    headers = {
        "Authorization": f"Bearer {secret_key}",
        "apikey": secret_key,
        "Content-Type": mime,
        "x-upsert": "false",
    }

    response = requests.post(
        endpoint,
        headers=headers,
        data=upload.getvalue(),
        timeout=30,
    )

    # If the exact name already exists (e.g. user retries after a network issue),
    # overwrite it safely rather than failing the whole inspection.
    if response.status_code == 400 and "already exists" in response.text.lower():
        headers["x-upsert"] = "true"
        response = requests.post(
            endpoint,
            headers=headers,
            data=upload.getvalue(),
            timeout=30,
        )

    if response.status_code not in (200, 201):
        raise RuntimeError(
            f"Photo upload failed ({response.status_code}). "
            "The inspection was not submitted; please try again."
        )

    # Store the private object path in PostgreSQL, never a temporary local path.
    return object_path


def signed_photo_url(object_path, expires_in=3600):
    """Create a short-lived URL for a private photo when the portal needs to display it."""
    if not object_path:
        return ""

    base_url, secret_key = _storage_config()

    from urllib.parse import quote
    encoded_path = quote(str(object_path).lstrip("/"), safe="/")
    endpoint = (
        f"{base_url}/storage/v1/object/sign/forklift-photos/{encoded_path}"
    )

    response = requests.post(
        endpoint,
        headers={
            "Authorization": f"Bearer {secret_key}",
            "apikey": secret_key,
            "Content-Type": "application/json",
        },
        json={"expiresIn": int(expires_in)},
        timeout=15,
    )

    if response.status_code not in (200, 201):
        return ""

    data = response.json()
    signed = data.get("signedURL") or data.get("signedUrl") or ""
    if not signed:
        return ""

    if signed.startswith("http"):
        return signed
    return f"{base_url}/storage/v1{signed}"

def smtp_ready():
    try:
        return all(
            st.secrets.get(k)
            for k in ["smtp_host","smtp_port","smtp_username","smtp_password","smtp_from"]
        )
    except Exception:
        return False

@st.cache_data(ttl=60, show_spinner=False)
def role_users(role, location=""):
    c = conn()
    rows = c.execute(
        "SELECT * FROM users WHERE role=? AND active=1 AND activated=1",
        (role,)
    ).fetchall()
    c.close()

    out = []
    for r in rows:
        rd = dict(r)
        if role == "WH Manager" and location:
            if rd.get("location") != location:
                continue
        out.append(rd)
    return out

def notify(case, role, subject, body, location=""):
    users = role_users(role, location)
    emails = [u["email"] for u in users if u.get("email")] or [""]

    c = conn()
    for email in emails:
        status = "Queued - email not configured"

        if email and smtp_ready():
            try:
                msg = EmailMessage()
                msg["Subject"] = subject
                msg["From"] = st.secrets["smtp_from"]
                msg["To"] = email
                msg.set_content(body)

                with smtplib.SMTP(
                    st.secrets["smtp_host"],
                    int(st.secrets["smtp_port"]),
                    timeout=15
                ) as s:
                    s.starttls()
                    s.login(
                        st.secrets["smtp_username"],
                        st.secrets["smtp_password"]
                    )
                    s.send_message(msg)

                status = "Sent"
            except Exception as ex:
                status = "Failed: " + str(ex)[:120]

        elif not email:
            status = "Queued - recipient email not set"

        c.execute("""
        INSERT INTO notifications(
            case_number,recipient_role,recipient_email,subject,body,status,created_at
        )
        VALUES(?,?,?,?,?,?,?)
        """, (case, role, email, subject, body, status, now()))

    c.commit()
    c.close()

def authenticate(username, password):
    username = (username or "").strip()

    c = conn()
    u = c.execute("""
        SELECT * FROM users
        WHERE LOWER(TRIM(username)) = LOWER(?)
          AND active = 1
          AND activated = 1
    """, (username,)).fetchone()
    c.close()

    if not u:
        return None

    if not u["password_salt"] or not u["password_hash"]:
        return None

    try:
        calculated_hash = hashlib.pbkdf2_hmac(
            "sha256",
            (password or "").encode("utf-8"),
            str(u["password_salt"]).encode("utf-8"),
            220000
        ).hex()

        if secrets.compare_digest(
            calculated_hash,
            str(u["password_hash"])
        ):
            return dict(u)

    except Exception:
        return None

    return None


def temporary_role_test_access():
    """TEMPORARY QA ACCESS - remove before production."""
    st.info("TEST MODE: Use this only to preview staff portals while the account workflow is being finalized.")

    with st.expander("🧪 Temporary Staff Portal Test Access", expanded=True):
        test_role = st.selectbox(
            "Preview Role",
            ["WH Manager", "Maintenance", "Compliance", "IT"],
            key="qa_preview_role"
        )

        test_name = st.text_input(
            "Test User Name",
            value="Portal Tester",
            key="qa_preview_name"
        )

        locations = sorted([x for x in equipment().location.unique() if x])

        if test_role == "WH Manager":
            test_location = st.selectbox(
                "Manager Location",
                locations,
                key="qa_preview_location"
            ) if locations else ""
        else:
            test_location = "ALL"
            st.caption("Access: All Locations")

        if st.button("OPEN TEST PORTAL", use_container_width=True, key="qa_open_portal"):
            st.session_state["user"] = {
                "id": -999,
                "username": "testpreview",
                "full_name": test_name.strip() or "Portal Tester",
                "role": test_role,
                "location": test_location,
                "email": "",
                "active": 1,
                "activated": 1,
                "must_change_password": 0,
                "test_mode": True,
            }
            st.rerun()


def staff_login():
    st.subheader("🔐 Staff Login")

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        ok = st.form_submit_button("LOGIN", use_container_width=True)

    if ok:
        u = authenticate(username, password)
        if u:
            st.session_state.user = u
            st.rerun()
        st.error("Invalid username or password.")

    st.divider()

    with st.expander("Activate my staff account / Set my password"):
        st.caption("Your role and location are assigned by IT. You set your own password.")

        with st.form("activate_form"):
            au = st.text_input("Username", key="au")
            code = st.text_input("Activation Code", key="ac")
            p1 = st.text_input("Create Password", type="password", key="ap1")
            p2 = st.text_input("Confirm Password", type="password", key="ap2")
            activate = st.form_submit_button("ACTIVATE ACCOUNT")

        if activate:
            if p1 != p2:
                st.error("Passwords do not match.")
            elif not strong_password(p1):
                st.error(
                    "Password must be 12+ characters and include uppercase, lowercase, number, and symbol."
                )
            else:
                clean_user = au.strip()
                clean_code = code.strip()

                c = conn()
                row = c.execute("""
                    SELECT * FROM users
                    WHERE LOWER(TRIM(username)) = LOWER(?)
                      AND TRIM(activation_code) = ?
                      AND active = 1
                """, (clean_user, clean_code)).fetchone()

                if not row:
                    c.close()
                    st.error("Invalid username or activation code.")
                else:
                    s, h = pw_hash(p1)

                    # Save the exact hash/salt format used by authenticate().
                    c.execute("""
                    UPDATE users
                    SET password_salt=?,
                        password_hash=?,
                        activated=1,
                        activation_code=NULL,
                        must_change_password=0,
                        temporary_password=FALSE,
                        password_changed_at=NOW()
                    WHERE id=?
                    """, (s, h, row["id"]))

                    c.commit()

                    # Immediately verify the password we just saved.
                    verify = c.execute("""
                        SELECT password_salt,password_hash,activated,active
                        FROM users WHERE id=?
                    """, (row["id"],)).fetchone()

                    test_hash = hashlib.pbkdf2_hmac(
                        "sha256",
                        p1.encode("utf-8"),
                        verify["password_salt"].encode("utf-8"),
                        220000
                    ).hex()

                    password_ok = secrets.compare_digest(
                        test_hash,
                        verify["password_hash"]
                    )

                    c.close()

                    if password_ok and int(verify["activated"]) == 1 and int(verify["active"]) == 1:
                        st.success("Account activated successfully. You can now log in with this username and password.")
                    else:
                        st.error("Activation saved, but password verification failed. Please contact IT.")

def record_last_login(user_id):
    c = conn()
    try:
        c.execute(
            "UPDATE users SET last_login_at=NOW(), updated_at=NOW() WHERE id=?",
            (int(user_id),)
        )
        c.commit()
    finally:
        c.close()


def force_password_change(u):
    if not u.get("must_change_password"):
        return False

    st.warning("You must create your own secure password before continuing.")

    with st.form("force_pw"):
        p1 = st.text_input("New Password", type="password")
        p2 = st.text_input("Confirm New Password", type="password")
        save = st.form_submit_button("SAVE NEW PASSWORD")

    if save:
        if p1 != p2:
            st.error("Passwords do not match.")
        elif not strong_password(p1):
            st.error(
                "Use 12+ characters with uppercase, lowercase, number, and symbol."
            )
        else:
            s, h = pw_hash(p1)
            c = conn()
            c.execute("""
            UPDATE users
            SET password_salt=?,
                password_hash=?,
                must_change_password=0,
                temporary_password=FALSE,
                password_changed_at=NOW()
            WHERE id=?
            """, (s, h, u["id"]))
            c.commit()
            c.close()

            st.session_state.user["must_change_password"] = 0
            st.success("Password changed.")
            st.rerun()

    return True

def logout_button():
    u = st.session_state.get("user")
    if u:
        loc = f" • {u['location']}" if u.get("location") and u["location"] != "ALL" else ""
        st.sidebar.success(f"{u['full_name']}\n\n{u['role']}{loc}")

        if st.sidebar.button("Logout", use_container_width=True):
            del st.session_state.user
            st.rerun()

def missing_today(location=None):
    eq = equipment()
    active = eq[~eq.status.str.lower().isin(["inactive","retired","deleted"])]

    if location and location != "ALL":
        active = active[active.location == location]

    done = query(
        "SELECT DISTINCT serial_number FROM inspections WHERE inspection_date=?",
        (str(date.today()),)
    )
    done_set = set(done.serial_number.astype(str)) if len(done) else set()
    return active[~active.serial_number.astype(str).isin(done_set)]

@st.cache_data(ttl=15, show_spinner=False)
def already_inspected(serial, shift):
    if not str(serial).strip() or not str(shift).strip(): return False
    c=conn()
    r=c.execute("""SELECT case_number,status FROM inspections
                   WHERE inspection_date=? AND UPPER(serial_number)=UPPER(?) AND shift=?
                   ORDER BY id DESC LIMIT 1""",
                (str(date.today()),str(serial),str(shift))).fetchone()
    c.close()
    return dict(r) if r else None

def humanize_columns(df):
    labels = {
        "case_number":"Case #","inspection_date":"Inspection Date","serial_number":"Serial #",
        "unit_number":"Unit #","operator_name":"Driver","driver":"Driver","brand":"Brand",
        "model":"Model","location":"Location","shift":"Shift","component":"Component","issue_type":"Issue",
        "driver_note":"Driver Note","created_at":"Reported At","inspection_time":"Inspection Time",
        "issue_reported":"Issue Reported","inspected_today":"Inspected Today","issue":"Issue",
        "status":"Status","out_of_service":"Out of Service","manager_name":"Manager",
        "manager_note":"Manager Note","maintenance_name":"Maintenance",
        "maintenance_note":"Maintenance Note","event_time":"Date / Time","actor_role":"Role",
        "actor_name":"Performed By","event_type":"Action","details":"Action / Notes",
        "found_location":"Reported Location"
    }
    return df.rename(columns={c: labels.get(c, c.replace("_"," ").title()) for c in df.columns})



def build_excel_bytes(sheets):
    """Create an .xlsx workbook without pandas ExcelWriter (Cloud-safe)."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    output = io.BytesIO()
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    added = False
    for sheet_name, df in sheets:
        safe_name = re.sub(r'[:\\\\/?*\\[\\]]', '-', str(sheet_name))[:31] or "Sheet"
        ws = wb.create_sheet(title=safe_name)
        added = True

        if df is None:
            df = pd.DataFrame()

        export_df = df.copy()
        export_df = export_df.where(pd.notna(export_df), None)

        if export_df.empty and len(export_df.columns) == 0:
            ws.append(["Message"])
            ws.append(["No records available for this report."])
        else:
            for row in dataframe_to_rows(export_df, index=False, header=True):
                ws.append([excel_safe_value(v) for v in list(row)])

            if export_df.empty:
                ws.append(["No records available."])

    if not added:
        ws = wb.create_sheet(title="No Data")
        ws.append(["Message"])
        ws.append(["No records available for this report."])

    wb.save(output)
    output.seek(0)
    return output.getvalue()

def excel_safe_value(value):
    """Convert pandas/numpy values into types openpyxl can safely serialize."""
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, pd.Timestamp):
        if value.tzinfo is not None:
            return value.strftime("%Y-%m-%d %H:%M:%S %Z")
        return value.to_pydatetime()

    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.strftime("%Y-%m-%d %H:%M:%S %Z")
        return value

    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass

    if isinstance(value, (list, dict, tuple, set)):
        return json.dumps(value, default=str)

    return value


def common_case_report(u):
    st.subheader("📋 Daily Logs & Reports")

    ins = query("SELECT * FROM inspections ORDER BY id DESC")

    # Managers only see their own warehouse everywhere in reports/history.
    if u["role"] == "WH Manager":
        ins = ins[ins.location == u["location"]]

    if ins.empty:
        st.info("No inspection records.")
        return

    available_dates = sorted(
        [d for d in ins.inspection_date.dropna().astype(str).unique() if d],
        reverse=True
    )

    selected_date = st.selectbox(
        "Select Date",
        available_dates,
        key=f"daily_log_date_{u['role']}_{u.get('location','ALL')}"
    )

    selected_shift = st.selectbox(
        "Shift",
        ["All Shifts", "Morning", "Afternoon", "Overnight"],
        key=f"daily_log_shift_{u['role']}_{u.get('location','ALL')}"
    )

    day = ins[ins.inspection_date.astype(str) == str(selected_date)].copy()

    if selected_shift != "All Shifts":
        day = day[day["shift"].fillna("") == selected_shift].copy()

    # Daily summary counts
    if u["role"] == "WH Manager":
        eq = equipment()
        active = eq[
            (eq.location == u["location"]) &
            (~eq.status.astype(str).str.lower().isin(["inactive","retired","deleted"]))
        ].copy()

        total_machines = len(active)

        if selected_shift == "All Shifts":
            expected_inspections = total_machines * 3
        else:
            expected_inspections = total_machines
    else:
        total_machines = day.serial_number.astype(str).str.upper().nunique() if len(day) else 0
        expected_inspections = "—"

    completed = len(day.drop_duplicates(subset=["serial_number","shift"])) if len(day) else 0
    issues_count = int(day.issue_found.fillna(0).astype(int).sum()) if len(day) else 0

    if u["role"] == "WH Manager":
        not_inspected = max(int(expected_inspections) - completed, 0)
    else:
        not_inspected = "—"

    a,b,c,d = st.columns(4)
    a.metric("Total Machines", total_machines)
    b.metric("Inspections Completed", completed)
    c.metric("Not Inspected", not_inspected)
    d.metric("Issues Reported", issues_count)

    # Shift summary for the selected day
    st.markdown("### Shift Summary")
    summary_rows = []
    for sh in ["Morning","Afternoon","Overnight"]:
        sh_day = ins[
            (ins.inspection_date.astype(str) == str(selected_date)) &
            (ins["shift"].fillna("") == sh)
        ].copy()

        sh_completed = len(sh_day.drop_duplicates(subset=["serial_number","shift"])) if len(sh_day) else 0
        sh_issues = int(sh_day.issue_found.fillna(0).astype(int).sum()) if len(sh_day) else 0

        if u["role"] == "WH Manager":
            sh_missing = max(total_machines - sh_completed, 0)
        else:
            sh_missing = "—"

        summary_rows.append({
            "Shift": sh,
            "Inspections Completed": sh_completed,
            "Not Inspected": sh_missing,
            "Issues Reported": sh_issues
        })

    st.dataframe(pd.DataFrame(summary_rows), hide_index=True, use_container_width=True)

    # Excel-only daily detail. Shift is included.
    log_cols = [
        "case_number","inspection_date","shift","location","serial_number","unit_number",
        "operator_name","issue_found","created_at"
    ]
    log_cols = [c for c in log_cols if c in day.columns]
    log = day[log_cols].copy()

    if "issue_found" in log.columns:
        log["issue_found"] = log["issue_found"].fillna(0).astype(int).map({0:"No",1:"Yes"})
        log = log.rename(columns={"issue_found":"issue_reported"})

    log = log.rename(columns={
        "operator_name":"driver",
        "created_at":"inspection_time"
    })

    issue_rows = query("SELECT * FROM issues")
    refs = day.case_number.tolist()
    excel_data = build_excel_bytes([
        ("Daily Inspection Log", humanize_columns(log)),
        ("Issues Reported", humanize_columns(issue_rows[issue_rows.case_number.isin(refs)])),
        ("Shift Summary", pd.DataFrame(summary_rows)),
    ])

    st.download_button(
        "⬇️ Download Daily Log",
        excel_data,
        f"forklift_daily_log_{selected_date}_{selected_shift.replace(' ','_')}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"download_daily_{u['role']}_{selected_date}_{selected_shift}"
    )

    st.divider()
    st.subheader("🕘 View Case History")
    st.caption("Issue cases only — follow each ticket from the driver's report until it is fixed.")

    issue_cases = ins[ins.issue_found.fillna(0).astype(int) == 1].copy()

    if issue_cases.empty:
        st.info("No issue cases are available.")
    else:
        h1,h2,h3,h4 = st.columns(4)

        if u["role"] == "WH Manager":
            selected_location = u["location"]
            h1.text_input("Location", value=u["location"], disabled=True)
        else:
            loc_options = ["All"] + sorted([x for x in issue_cases.location.dropna().unique() if x])
            selected_location = h1.selectbox(
                "Location",
                loc_options,
                key=f"history_location_{u['role']}"
            )

        selected_history_shift = h2.selectbox(
            "Shift",
            ["All Shifts","Morning","Afternoon","Overnight"],
            key=f"history_shift_{u['role']}"
        )

        status_options = ["All"] + sorted([x for x in issue_cases.status.dropna().unique() if x])
        selected_status = h3.selectbox(
            "Status",
            status_options,
            key=f"history_status_{u['role']}"
        )

        oos = h4.selectbox(
            "Out of Service",
            ["All","Yes","No"],
            key=f"history_oos_{u['role']}"
        )

        hist = issue_cases.copy()

        if u["role"] != "WH Manager" and selected_location != "All":
            hist = hist[hist.location == selected_location]

        if selected_history_shift != "All Shifts":
            hist = hist[hist["shift"].fillna("") == selected_history_shift]

        if selected_status != "All":
            hist = hist[hist.status == selected_status]

        if oos != "All":
            hist = hist[hist.out_of_service.fillna("No") == oos]

        search = st.text_input(
            "Search Case # / Serial # / Unit # / Driver",
            key=f"history_search_{u['role']}"
        ).strip().lower()

        if search:
            hist = hist[
                hist.case_number.astype(str).str.lower().str.contains(search, na=False) |
                hist.serial_number.astype(str).str.lower().str.contains(search, na=False) |
                hist.unit_number.astype(str).str.lower().str.contains(search, na=False) |
                hist.operator_name.astype(str).str.lower().str.contains(search, na=False)
            ]

        history_cols = [
            "case_number","inspection_date","shift","location","serial_number","unit_number",
            "operator_name","status","out_of_service"
        ]

        st.dataframe(
            humanize_columns(hist[[c for c in history_cols if c in hist.columns]]),
            hide_index=True,
            use_container_width=True
        )

    st.divider()
    st.subheader("🚜 Machine History")

    # Machine History uses the asset master so machines with zero tickets are also visible.
    eq = equipment()
    active_eq = eq[
        ~eq.status.astype(str).str.lower().isin(["inactive","retired","deleted"])
    ].copy()

    if u["role"] == "WH Manager":
        active_eq = active_eq[active_eq.location == u["location"]].copy()
        st.caption(f"All active machines at {u['location']} — including machines with no tickets.")
    else:
        st.caption("All active machines — including machines with no tickets.")

    if active_eq.empty:
        st.info("No active machines available.")
        return

    machine_labels = []
    machine_lookup = {}
    for _, er in active_eq.iterrows():
        label = f"{er.serial_number} • Unit {er.unit_number or '—'} • {er.brand} • {er.location}"
        machine_labels.append(label)
        machine_lookup[label] = er

    selected_machine = st.selectbox(
        "Select Machine",
        machine_labels,
        index=None,
        placeholder="Choose a machine",
        key=f"machine_history_select_{u['role']}_{u.get('location','ALL')}"
    )

    if selected_machine:
        er = machine_lookup[selected_machine]
        serial = str(er.serial_number)

        mh_all = ins[ins.serial_number.astype(str).str.upper() == serial.upper()].copy()
        mh_issues = mh_all[mh_all.issue_found.fillna(0).astype(int) == 1].copy()

        fixed_mask = mh_issues.status.fillna("").astype(str).str.startswith("Fixed by")

        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Total Inspections", len(mh_all))
        m2.metric("Total Tickets", len(mh_issues))
        m3.metric("Fixed Tickets", int(fixed_mask.sum()))
        m4.metric("Open Tickets", int((~fixed_mask).sum()))

        st.markdown(
            f"**Serial:** {er.serial_number}  •  "
            f"**Unit:** {er.unit_number or '—'}  •  "
            f"**Brand:** {er.brand or '—'}  •  "
            f"**Model:** {er.model or '—'}  •  "
            f"**Location:** {er.location}"
        )

        # Show full inspection history, including no-issue inspections.
        st.markdown("#### Inspection History")
        if mh_all.empty:
            st.info("No inspections have been recorded for this machine yet.")
        else:
            inspect_cols = [
                "inspection_date","shift","case_number","operator_name","issue_found","status"
            ]
            mh_inspect = mh_all[[c for c in inspect_cols if c in mh_all.columns]].copy()
            if "issue_found" in mh_inspect.columns:
                mh_inspect["issue_found"] = mh_inspect["issue_found"].fillna(0).astype(int).map({0:"No",1:"Yes"})
                mh_inspect = mh_inspect.rename(columns={"issue_found":"issue"})
            st.dataframe(
                humanize_columns(mh_inspect),
                hide_index=True,
                use_container_width=True
            )

        st.markdown("#### Ticket History")
        if mh_issues.empty:
            st.success("No issue tickets have been recorded for this machine.")
        else:
            rows = []
            for _, mr in mh_issues.iterrows():
                mi = query(
                    "SELECT component,issue_type FROM issues WHERE case_number=?",
                    (mr.case_number,)
                )
                issue_text = "; ".join(
                    [f"{x.component} — {x.issue_type}" for _, x in mi.iterrows()]
                )

                rows.append({
                    "inspection_date": mr.inspection_date,
                    "shift": mr.get("shift",""),
                    "location": mr.location,
                    "case_number": mr.case_number,
                    "operator_name": mr.operator_name,
                    "issue": issue_text,
                    "status": mr.status
                })

            st.dataframe(
                humanize_columns(pd.DataFrame(rows)),
                hide_index=True,
                use_container_width=True
            )

    # Case detail stays below Machine History.
    if issue_cases.empty:
        return

    if 'hist' in locals() and not hist.empty:
        selected_case = st.selectbox(
            "Select Issue Case",
            hist.case_number.tolist(),
            key=f"history_case_{u['role']}"
        )

        rec = hist[hist.case_number == selected_case].iloc[0]

        st.markdown(
            f"**Case {selected_case}** • "
            f"Shift **{rec.get('shift','—') or '—'}** • "
            f"Serial **{rec.serial_number}** • "
            f"Driver **{rec.operator_name}** • "
            f"Status **{rec.status}**"
        )

        ii = query(
            "SELECT component,issue_type,driver_note,created_at FROM issues WHERE case_number=?",
            (selected_case,)
        )

        if not ii.empty:
            st.markdown("**Reported Issue(s)**")
            st.dataframe(
                humanize_columns(ii),
                hide_index=True,
                use_container_width=True
            )

        timeline = query(
            """SELECT event_time,actor_role,actor_name,event_type,details
               FROM audit
               WHERE case_number=?
               ORDER BY event_time""",
            (selected_case,)
        )

        if not timeline.empty:
            reported_issues = query(
                "SELECT component,issue_type,driver_note FROM issues WHERE case_number=?",
                (selected_case,)
            )

            readable = []
            for _, tr in timeline.iterrows():
                event = str(tr.event_type or "")
                raw = str(tr.details or "").strip()

                if event == "Inspection Submitted":
                    notes = [
                        str(x).strip()
                        for x in reported_issues.driver_note.tolist()
                        if str(x or "").strip()
                    ]
                    detail = "; ".join(notes) if notes else "—"

                elif event in ("Need More Time","Repair Needs More Time"):
                    detail = raw.split("; ",1)[1] if "; " in raw else (raw or "—")

                else:
                    detail = raw if raw else "—"

                readable.append(detail)

            timeline = timeline.copy()
            timeline["details"] = readable

            st.markdown("**Case Timeline — Full Lifecycle**")
            st.dataframe(
                humanize_columns(timeline),
                hide_index=True,
                use_container_width=True
            )

def run_driver():
    # Clear all previous inspection widget values after a successful submit.
    if st.session_state.pop("drv_reset_after_submit", False):
        for key in list(st.session_state.keys()):
            if str(key).startswith("drv_"):
                st.session_state.pop(key, None)

    st.header("🚜 Daily Forklift Inspection")
    st.caption("No login required • Every active forklift must be inspected every day.")

    eq = equipment()
    active = eq[~eq.status.str.lower().isin(["inactive","retired","deleted"])]
    locs = sorted([x for x in active.location.unique() if x])

    if not locs:
        st.error("No active locations/assets are available.")
        return

    c1, c2, c3 = st.columns(3)
    loc = c1.selectbox("📍 Location", locs, key="drv_location")
    shift = c2.selectbox("🕒 Shift *", ["Morning","Afternoon","Overnight"], index=None,
                         placeholder="Choose shift", key="drv_shift")
    local = active[active.location == loc]

    labels = []
    lookup = {}

    for _, r in local.iterrows():
        oos = " • OUT OF SERVICE" if str(r.out_of_service).lower() == "yes" else ""
        label = (
            f"{r.serial_number or 'NO SERIAL'} • "
            f"Unit {r.unit_number or '—'} • {r.brand}{oos}"
        )
        labels.append(label)
        lookup[label] = r

    labels.append("⚠️ MACHINE NOT LISTED")
    choice = c3.selectbox("🔎 Machine / Serial Number", labels, key="drv_machine")

    unlisted = choice == "⚠️ MACHINE NOT LISTED"

    if unlisted:
        st.warning(
            "Machine is not listed at this location. Enter the information visible on the machine."
        )
        a, b, c, d = st.columns(4)
        serial = a.text_input("Serial Number *", key="drv_serial")
        unit = b.text_input("Unit / Truck #", key="drv_unit")
        brand = c.text_input("Brand (if known)", key="drv_brand")
        model = d.text_input("Model (if known)", key="drv_model")
    else:
        r = lookup[choice]
        serial = r.serial_number
        unit = r.unit_number
        brand = r.brand
        model = r.model

        dup = already_inspected(serial, shift)

        st.markdown(
            f"<div class='machine-selected'>"
            f"Selected machine: <b>Serial {serial or 'Not recorded'}</b> • "
            f"Unit {unit or '—'} • {r.brand} {r.model}"
            f"</div>",
            unsafe_allow_html=True
        )

        if dup:
            st.warning(
                f"This machine was already inspected for the **{shift} shift** today under case "
                f"**{dup['case_number']}**. Another inspection for the same shift cannot be submitted."
            )

    operator = st.text_input("👤 Inspector / Operator Name *", key="drv_operator")

    rows = []
    results = {}
    missing_sections = []
    issue_validation = []

    st.divider()

    for gi, (group, items) in enumerate(GROUPS):
        with st.expander(group, expanded=(gi == 0)):
            st.write("Inspect every item below before choosing the group result:")
            st.write(" • " + "  \n • ".join(items))

            ans = st.radio(
                "Group Result",
                ["✅ All listed items inspected — OK", "⚠️ Issue Found"],
                index=None,
                horizontal=True,
                key=f"drv_group_{gi}"
            )

            results[group] = ans

            if ans is None:
                # Keep enough detail in the error so the driver knows exactly where to go.
                clean_group = re.sub(r"^[^A-Za-z0-9]+", "", group).strip()
                important_items = ", ".join(items[:4])
                if "Horn" in items:
                    important_items += ", Horn"
                missing_sections.append(
                    f"**{clean_group}** — no OK / Issue selection was made "
                    f"(items include: {important_items})."
                )

            if ans == "⚠️ Issue Found":
                comps = st.multiselect(
                    "Which component has an issue?",
                    items,
                    key=f"drv_comp_{gi}"
                )

                if not comps:
                    clean_group = re.sub(r"^[^A-Za-z0-9]+", "", group).strip()
                    issue_validation.append(
                        f"**{clean_group}** — select the component with the issue."
                    )

                for j, comp in enumerate(comps):
                    typ = st.selectbox(
                        f"{comp} — Issue Type",
                        ISSUE_TYPES.get(comp, generic_issue_types(comp)),
                        key=f"drv_type_{gi}_{j}"
                    )

                    txt = st.text_area(
                        f"{comp} — What did you notice? *",
                        placeholder="Briefly describe what you saw, heard, felt, or noticed.",
                        key=f"drv_note_{gi}_{j}"
                    )

                    pic = st.file_uploader(
                        f"{comp} — Photo (optional)",
                        type=["jpg","jpeg","png"],
                        key=f"drv_photo_{gi}_{j}"
                    )

                    if not txt.strip():
                        issue_validation.append(
                            f"**{comp}** — 'What did you notice?' is required."
                        )

                    rows.append((group, comp, typ, txt, pic))

    severity = (
        st.selectbox(
            "Severity",
            ["Minor","Operational","Safety Critical"],
            key="drv_severity"
        )
        if rows else ""
    )

    confirm = st.checkbox(
        "I confirm that I physically inspected all listed items and the information is accurate.",
        key="drv_confirm"
    )

    submit_clicked = st.button(
        "SUBMIT INSPECTION",
        type="primary",
        use_container_width=True,
        key="drv_submit"
    )

    # Keep confirmation directly below the Submit button.
    success_message = st.session_state.pop("inspection_success_message", None)
    if success_message:
        st.success(success_message, icon="✅")

    if submit_clicked:
        errors = []

        if not operator.strip():
            errors.append("**Inspector / Operator Name** is blank.")

        if unlisted and not serial.strip():
            errors.append("**Machine Not Listed → Serial Number** is blank.")

        if not unlisted and already_inspected(serial, shift):
            errors.append(f"Serial **{serial}** has already been inspected today.")

        errors.extend(missing_sections)
        errors.extend(issue_validation)

        if not confirm:
            errors.append("**Inspection confirmation** is not checked.")

        if errors:
            st.error(
                "❌ Inspection was NOT submitted. Complete the items below and submit again:"
            )
            for e in errors:
                st.markdown(f"- {e}")
            return

        case = daily_case_number()
        status = (
            "Not Reviewed by Manager"
            if rows
            else "Inspected"
        )

        c = conn()
        c.execute("""
        INSERT INTO inspections(
            case_number,inspection_date,serial_number,unit_number,location,shift,
            operator_name,machine_listed,group_results,issue_found,severity,
            status,created_at,updated_at
        )
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            case, str(date.today()), serial, unit, loc, shift, operator.strip(),
            0 if unlisted else 1, json.dumps(results),
            1 if rows else 0, severity, status, now(), now()
        ))

        for i, (g, comp, typ, txt, pic) in enumerate(rows, 1):
            pp = save_photo(
                pic,
                f"{case}_{i}_{re.sub('[^A-Za-z0-9]','',comp)}"
            )
            c.execute("""
            INSERT INTO issues(
                case_number,group_name,component,issue_type,
                driver_note,photo_path,created_at
            )
            VALUES(?,?,?,?,?,?,?)
            """, (case, g, comp, typ, txt, pp, now()))

        if unlisted:
            c.execute("""
            INSERT INTO asset_review(
                case_number,serial_number,found_location,brand,model,unit_number,
                status,updated_at
            )
            VALUES(?,?,?,?,?,?,'Waiting Compliance Review',?)
            """, (case, serial, loc, brand, model, unit, now()))

        c.commit()
        c.close()

        already_inspected.clear()

        audit(
            case,
            "Driver",
            operator.strip(),
            "Inspection Submitted",
            f"Serial={serial}; Issues={len(rows)}; Machine Not Listed={unlisted}"
        )

        notify(
            case,
            "WH Manager",
            f"Forklift inspection requires review - {case}",
            f"Case {case}\nSerial: {serial}\nLocation: {loc}\n"
            f"Inspector: {operator.strip()}\nOpen your Manager Portal to review.",
            loc
        )

        if unlisted:
            notify(
                case,
                "Maintenance",
                f"Unlisted machine reported - {case}",
                f"Case {case}\nSerial: {serial}\nFound at: {loc}\n"
                f"Driver: {operator.strip()}."
            )
            notify(
                case,
                "Compliance",
                f"Unlisted machine reported - {case}",
                f"Case {case}\nSerial: {serial}\nFound at: {loc}\n"
                f"Driver: {operator.strip()}."
            )

        st.session_state["inspection_success_message"] = (
            f"Inspection has been submitted successfully — Serial: {serial}"
        )
        st.session_state["drv_reset_after_submit"] = True
        st.rerun()



def daily_report_tab(u, key_prefix, all_locations=False):
    st.subheader("📋 Daily Logs & Reports")

    if u["role"] == "WH Manager":
        date_rows = query(
            """SELECT DISTINCT inspection_date
               FROM inspections
               WHERE location=?
               ORDER BY inspection_date DESC""",
            (u["location"],)
        )
    else:
        date_rows = query(
            """SELECT DISTINCT inspection_date
               FROM inspections
               ORDER BY inspection_date DESC"""
        )

    if date_rows.empty:
        st.info("No inspection records.")
        return

    dates = [str(d) for d in date_rows.inspection_date.dropna().tolist()]

    cols = st.columns(3 if all_locations else 2)

    selected_date = cols[0].selectbox(
        "Date",
        dates,
        key=f"{key_prefix}_date"
    )

    if all_locations:
        locations = ["All"] + sorted(
            [x for x in equipment().location.dropna().astype(str).unique() if x]
        )
        selected_location = cols[1].selectbox(
            "Location",
            locations,
            key=f"{key_prefix}_location"
        )
        selected_shift = cols[2].selectbox(
            "Shift",
            ["All Shifts","Morning","Afternoon","Overnight"],
            key=f"{key_prefix}_shift"
        )
    else:
        selected_location = u["location"]
        selected_shift = cols[1].selectbox(
            "Shift",
            ["All Shifts","Morning","Afternoon","Overnight"],
            key=f"{key_prefix}_shift"
        )
        st.caption(f"Location: {u['location']}")

    sql = "SELECT * FROM inspections WHERE inspection_date=?"
    params = [str(selected_date)]

    if selected_location != "All":
        sql += " AND location=?"
        params.append(selected_location)

    if selected_shift != "All Shifts":
        sql += " AND shift=?"
        params.append(selected_shift)

    sql += " ORDER BY id DESC"
    day = query(sql, tuple(params))

    completed = len(
        day.drop_duplicates(subset=["serial_number","shift"])
    ) if len(day) else 0

    issues_reported = int(
        day.issue_found.fillna(0).astype(int).sum()
    ) if len(day) else 0

    if u["role"] == "WH Manager":
        eq = equipment()
        active = eq[
            (eq.location == u["location"]) &
            (~eq.status.astype(str).str.lower().isin(["inactive","retired","deleted"]))
        ].copy()
        total_machines = len(active)
        expected = total_machines * (3 if selected_shift == "All Shifts" else 1)
        not_inspected = max(expected - completed, 0)

        a,b,c,d = st.columns(4)
        a.metric("Total Machines", total_machines)
        b.metric("Inspections Completed", completed)
        c.metric("Not Inspected", not_inspected)
        d.metric("Issues Reported", issues_reported)
    else:
        repair_not_started = len(day[day.status == "Repair Not Started"])
        fixed = len(day[day.status.fillna("").astype(str).str.startswith("Fixed by")])

        a,b,c,d = st.columns(4)
        a.metric("Inspections Completed", completed)
        b.metric("Issues Reported", issues_reported)
        c.metric("Repair Not Started", repair_not_started)
        d.metric("Fixed", fixed)

    st.markdown("### Shift Summary")
    base_sql = "SELECT * FROM inspections WHERE inspection_date=?"
    base_params = [str(selected_date)]
    if selected_location != "All":
        base_sql += " AND location=?"
        base_params.append(selected_location)
    base = query(base_sql + " ORDER BY id DESC", tuple(base_params))

    summary_rows = []
    for sh in ["Morning","Afternoon","Overnight"]:
        sh_day = base[base["shift"].fillna("") == sh].copy()
        row = {
            "Shift": sh,
            "Inspections Completed": len(sh_day.drop_duplicates(subset=["serial_number","shift"])) if len(sh_day) else 0,
            "Issues Reported": int(sh_day.issue_found.fillna(0).astype(int).sum()) if len(sh_day) else 0,
        }
        if u["role"] == "WH Manager":
            row["Not Inspected"] = max(total_machines - row["Inspections Completed"], 0)
        else:
            row["Repair Not Started"] = len(sh_day[sh_day.status == "Repair Not Started"])
            row["Fixed"] = len(sh_day[sh_day.status.fillna("").astype(str).str.startswith("Fixed by")])
        summary_rows.append(row)

    st.dataframe(pd.DataFrame(summary_rows), hide_index=True, use_container_width=True)

    log_cols = [
        "case_number","inspection_date","shift","location","serial_number",
        "unit_number","operator_name","issue_found","status","created_at"
    ]
    log = day[[c for c in log_cols if c in day.columns]].copy()
    if "issue_found" in log.columns:
        log["issue_found"] = log["issue_found"].fillna(0).astype(int).map({0:"No",1:"Yes"})
        log = log.rename(columns={"issue_found":"issue_reported"})
    log = log.rename(columns={"operator_name":"driver","created_at":"inspection_time"})

    excel_data = build_excel_bytes([
        ("Daily Log", humanize_columns(log)),
        ("Shift Summary", pd.DataFrame(summary_rows)),
    ])

    st.download_button(
        "⬇️ Download Daily Log",
        excel_data,
        f"{key_prefix}_daily_log_{selected_date}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"{key_prefix}_download"
    )


def case_machine_history_tab(u, key_prefix, all_locations=False):
    st.subheader("🕘 View Case History")

    if u["role"] == "WH Manager":
        issue_cases = query(
            """SELECT * FROM inspections
               WHERE location=? AND issue_found=1
               ORDER BY id DESC""",
            (u["location"],)
        )
    else:
        issue_cases = query(
            """SELECT * FROM inspections
               WHERE issue_found=1
               ORDER BY id DESC"""
        )

    if issue_cases.empty:
        st.info("No issue cases are available.")
        hist = pd.DataFrame()
    else:
        cols = st.columns(4)

        if all_locations:
            location = cols[0].selectbox(
                "Location",
                ["All"] + sorted([x for x in issue_cases.location.dropna().astype(str).unique() if x]),
                key=f"{key_prefix}_history_location"
            )
        else:
            location = u["location"]
            cols[0].text_input("Location", value=u["location"], disabled=True, key=f"{key_prefix}_history_location")

        shift = cols[1].selectbox(
            "Shift",
            ["All Shifts","Morning","Afternoon","Overnight"],
            key=f"{key_prefix}_history_shift"
        )
        status = cols[2].selectbox(
            "Status",
            ["All"] + sorted([x for x in issue_cases.status.dropna().astype(str).unique() if x]),
            key=f"{key_prefix}_history_status"
        )
        oos = cols[3].selectbox(
            "Out of Service",
            ["All","Yes","No"],
            key=f"{key_prefix}_history_oos"
        )

        hist = issue_cases.copy()
        if location != "All":
            hist = hist[hist.location == location]
        if shift != "All Shifts":
            hist = hist[hist["shift"].fillna("") == shift]
        if status != "All":
            hist = hist[hist.status == status]
        if oos != "All":
            hist = hist[hist.out_of_service.fillna("No") == oos]

        search = st.text_input(
            "Search Case # / Serial # / Unit # / Driver",
            key=f"{key_prefix}_history_search"
        ).strip().lower()

        if search:
            hist = hist[
                hist.case_number.astype(str).str.lower().str.contains(search, na=False) |
                hist.serial_number.astype(str).str.lower().str.contains(search, na=False) |
                hist.unit_number.astype(str).str.lower().str.contains(search, na=False) |
                hist.operator_name.astype(str).str.lower().str.contains(search, na=False)
            ]

        cols_show = [
            "case_number","inspection_date","shift","location","serial_number",
            "unit_number","operator_name","status","out_of_service"
        ]
        st.dataframe(
            humanize_columns(hist[[c for c in cols_show if c in hist.columns]]),
            hide_index=True,
            use_container_width=True
        )

        if not hist.empty:
            selected_case = st.selectbox(
                "Select Issue Case",
                hist.case_number.tolist(),
                key=f"{key_prefix}_history_case"
            )
            rec = hist[hist.case_number == selected_case].iloc[0]

            st.markdown(
                f"**Case {selected_case}** • "
                f"Shift **{rec.get('shift','—') or '—'}** • "
                f"Serial **{rec.serial_number}** • "
                f"Driver **{rec.operator_name}** • "
                f"Status **{rec.status}**"
            )

            ii = query(
                "SELECT component,issue_type,driver_note,created_at FROM issues WHERE case_number=?",
                (selected_case,)
            )
            if not ii.empty:
                st.markdown("**Reported Issue(s)**")
                st.dataframe(humanize_columns(ii), hide_index=True, use_container_width=True)

            timeline = query(
                """SELECT event_time,actor_role,actor_name,event_type,details
                   FROM audit WHERE case_number=? ORDER BY event_time""",
                (selected_case,)
            )
            if not timeline.empty:
                driver_notes = query(
                    "SELECT driver_note FROM issues WHERE case_number=?",
                    (selected_case,)
                )
                readable = []
                for _, tr in timeline.iterrows():
                    event = str(tr.event_type or "")
                    raw = str(tr.details or "").strip()
                    if event == "Inspection Submitted":
                        notes = [str(x).strip() for x in driver_notes.driver_note.tolist() if str(x or "").strip()]
                        detail = "; ".join(notes) if notes else "—"
                    elif event in ("Need More Time","Repair Needs More Time"):
                        detail = raw.split("; ",1)[1] if "; " in raw else (raw or "—")
                    else:
                        detail = raw if raw else "—"
                    readable.append(detail)

                timeline = timeline.copy()
                timeline["details"] = readable
                st.markdown("**Case Timeline — Full Lifecycle**")
                st.dataframe(humanize_columns(timeline), hide_index=True, use_container_width=True)

    st.divider()
    st.subheader("🚜 Machine History")

    eq = equipment()
    active_eq = eq[
        ~eq.status.astype(str).str.lower().isin(["inactive","retired","deleted"])
    ].copy()

    if u["role"] == "WH Manager":
        active_eq = active_eq[active_eq.location == u["location"]]
        st.caption(f"All active machines at {u['location']} — including machines with no tickets.")
    elif all_locations:
        mh_loc = st.selectbox(
            "Machine Location",
            ["All"] + sorted([x for x in active_eq.location.dropna().astype(str).unique() if x]),
            key=f"{key_prefix}_machine_location"
        )
        if mh_loc != "All":
            active_eq = active_eq[active_eq.location == mh_loc]
        st.caption("All active machines — including machines with no tickets.")

    if active_eq.empty:
        st.info("No active machines available.")
        return

    machine_labels = []
    machine_lookup = {}
    for _, er in active_eq.iterrows():
        label = f"{er.serial_number} • Unit {er.unit_number or '—'} • {er.brand} • {er.location}"
        machine_labels.append(label)
        machine_lookup[label] = er

    selected_machine = st.selectbox(
        "Select Machine",
        machine_labels,
        index=None,
        placeholder="Choose a machine",
        key=f"{key_prefix}_machine_select"
    )

    if selected_machine:
        er = machine_lookup[selected_machine]
        serial = str(er.serial_number)

        # Full lifetime history is intentionally NOT restricted to the manager's
        # current warehouse. The machine selector is location-scoped, but once a
        # serial is selected we show every inspection ever recorded for that serial.
        mh_all = query(
            """SELECT * FROM inspections
               WHERE UPPER(serial_number)=UPPER(?)
               ORDER BY inspection_date DESC, created_at DESC, id DESC""",
            (serial,)
        )
        mh_issues = mh_all[mh_all.issue_found.fillna(0).astype(int) == 1].copy()
        fixed = mh_issues.status.fillna("").astype(str).str.startswith("Fixed by")

        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Driver Inspections Submitted", len(mh_all))
        m2.metric("Total Tickets", len(mh_issues))
        m3.metric("Fixed Tickets", int(fixed.sum()))
        m4.metric("Open Tickets", int((~fixed).sum()))

        st.markdown(
            f"**Serial:** {er.serial_number} • **Unit:** {er.unit_number or '—'} • "
            f"**Brand:** {er.brand or '—'} • **Model:** {er.model or '—'} • "
            f"**Location:** {er.location}"
        )

        st.markdown("#### Inspection History")
        if mh_all.empty:
            st.info("No inspections have been recorded for this machine.")
        else:
            cols = ["inspection_date","shift","location","case_number","operator_name","issue_found","status"]
            d = mh_all[[c for c in cols if c in mh_all.columns]].copy()
            d["issue_found"] = d["issue_found"].fillna(0).astype(int).map({0:"No",1:"Yes"})
            d = d.rename(columns={"issue_found":"issue"})
            st.dataframe(humanize_columns(d), hide_index=True, use_container_width=True)

        st.markdown("#### Ticket History")
        if mh_issues.empty:
            st.success("No issue tickets have been recorded for this machine.")
        else:
            rows = []
            for _, mr in mh_issues.iterrows():
                mi = query("SELECT component,issue_type FROM issues WHERE case_number=?", (mr.case_number,))
                issue_text = "; ".join([f"{x.component} — {x.issue_type}" for _, x in mi.iterrows()])
                rows.append({
                    "inspection_date": mr.inspection_date,
                    "shift": mr.get("shift",""),
                    "case_number": mr.case_number,
                    "operator_name": mr.operator_name,
                    "issue": issue_text,
                    "status": mr.status
                })
            st.dataframe(humanize_columns(pd.DataFrame(rows)), hide_index=True, use_container_width=True)

def manager_actions(u):
    if not u.get("location") or u["location"] == "ALL":
        st.error("IT must assign this Manager to one location.")
        return

    hour = datetime.now().hour
    greeting = "Good Morning" if hour < 12 else ("Good Afternoon" if hour < 17 else "Good Evening")
    st.header(f"{greeting}, {u['full_name']} 👋")
    st.caption(u["location"])

    tab_work, tab_reports, tab_history = st.tabs(
        ["📌 My Daily Work", "📋 Reports", "🕘 Case & Machine History"]
    )

    with tab_work:
        today_all = query(
            """SELECT * FROM inspections
               WHERE location=? AND inspection_date=?
               ORDER BY id DESC""",
            (u["location"], str(date.today()))
        )
        shift_view=st.segmented_control("Shift",["Morning","Afternoon","Overnight","All Shifts"],
                                        default="Morning",key="manager_shift_view")
        today=today_all if shift_view=="All Shifts" else today_all[today_all["shift"].fillna("")==shift_view].copy()
        issue_review=today[today.status=="Not Reviewed by Manager"].copy()

        eq = equipment()
        active = eq[
            (eq.location == u["location"]) &
            (~eq.status.astype(str).str.lower().isin(["inactive","retired","deleted"]))
        ].copy()

        inspected_unique = today.serial_number.astype(str).str.upper().nunique() if len(today) else 0
        issues_today = int(today.issue_found.fillna(0).astype(int).sum()) if len(today) else 0
        expected=len(active)*(3 if shift_view=="All Shifts" else 1)
        completed=len(today.drop_duplicates(subset=["serial_number","shift"])) if len(today) else 0
        not_inspected=max(expected-completed,0)

        st.markdown("### Today's Overview")
        a,b,c,d = st.columns(4)
        a.metric("Total Machines", len(active))
        b.metric("Inspections Completed", completed)
        c.metric("Not Inspected Today", not_inspected)
        d.metric("Issues Found Today", issues_today)

        # One clear daily machine table: every machine + inspected yes/no + driver.
        latest_today = today.sort_values("id").drop_duplicates(
            subset=["serial_number","shift"], keep="last"
        ) if len(today) else pd.DataFrame()

        driver_map = {}
        case_map = {}
        issue_map = {}
        time_map = {}
        if len(latest_today):
            for _, r in latest_today.iterrows():
                key = str(r.serial_number).strip().upper()
                driver_map[key] = r.operator_name
                case_map[key] = r.case_number
                issue_map[key] = "Yes" if int(r.issue_found or 0) else "No"
                time_map[key] = r.created_at

        machine_rows=[]
        shifts_to_show=["Morning","Afternoon","Overnight"] if shift_view=="All Shifts" else [shift_view]
        for _,r in active.iterrows():
            skey=str(r.serial_number).strip().upper()
            for sh in shifts_to_show:
                match=latest_today[(latest_today.serial_number.astype(str).str.strip().str.upper()==skey)&
                                   (latest_today["shift"].fillna("")==sh)] if len(latest_today) else pd.DataFrame()
                rr=match.iloc[-1] if len(match) else None
                machine_rows.append({"serial_number":r.serial_number,"unit_number":r.unit_number,
                    "brand":r.brand,"model":r.model,"shift":sh,
                    "inspected_today":"Yes" if rr is not None else "No",
                    "driver":rr.operator_name if rr is not None else "—",
                    "issue":("Yes" if int(rr.issue_found or 0) else "No") if rr is not None else "—",
                    "inspection_time":rr.created_at if rr is not None else "—"})

        machine_status = pd.DataFrame(machine_rows)
        with st.expander(f"Today's Inspection Status — {len(active)} machine(s)", expanded=False):
            st.dataframe(humanize_columns(machine_status), hide_index=True, use_container_width=True)

            status_buf = io.BytesIO()
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws = wb.active
            ws.title = "Inspection Status"

            _export_df = humanize_columns(machine_status.copy())
            if _export_df.empty:
                ws.append(["Message"])
                ws.append(["No inspection status records are available."])
            else:
                for _row in dataframe_to_rows(_export_df, index=False, header=True):
                    ws.append([excel_safe_value(v) for v in list(_row)])

            wb.save(status_buf)
            status_buf.seek(0)
            st.download_button(
                "⬇️ Download Today's Inspection Status",
                status_buf.getvalue(),
                f"inspection_status_{u['location']}_{date.today()}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"download_today_status_{u['location']}"
            )

        st.markdown("### My Today's Tasks")
        st.caption("Only inspections with an issue that still need your decision.")

        if issue_review.empty:
            st.success("No issue inspections are waiting for your review.")
        else:
            for _, r in issue_review.iterrows():
                with st.expander(
                    f"{r.case_number} • Serial {r.serial_number} • Driver {r.operator_name}",
                    expanded=False
                ):
                    st.caption(f"Unit {r.unit_number or '—'} • Submitted {r.created_at}")
                    issues = query(
                        "SELECT component,issue_type,driver_note FROM issues WHERE case_number=?",
                        (r.case_number,)
                    )
                    st.dataframe(humanize_columns(issues), hide_index=True, use_container_width=True)

                    action = st.radio(
                        "What do you want to do?",
                        ["✅ FIXED", "🔧 SEND TO MAINTENANCE"],
                        horizontal=True,
                        key=f"mgr_action_{r.id}"
                    )
                    manager_name=st.text_input("Manager Name *",key=f"mgr_name_{r.id}")
                    oos = st.radio(
                        "Out of Service?",
                        ["No","Yes"],
                        horizontal=True,
                        key=f"mgr_oos_{r.id}"
                    )
                    note = st.text_area(
                        "Manager note *",
                        key=f"mgr_note_{r.id}"
                    )

                    if st.button("SAVE", key=f"mgr_save_{r.id}", type="primary"):
                        if not manager_name.strip():
                            st.error("Manager Name is required.")
                        elif not note.strip():
                            st.error("Manager note is required.")
                        else:
                            sent = action.startswith("🔧")
                            status = "Repair Not Started" if sent else f"Fixed by {manager_name.strip()}"

                            c = conn()
                            c.execute(
                                """UPDATE inspections
                                   SET manager_name=?,manager_decision=?,manager_note=?,manager_at=?,
                                       out_of_service=?,status=?,closed_at=?,updated_at=?
                                   WHERE id=?""",
                                (
                                    manager_name.strip(),
                                    "Repair Required" if sent else "Fixed by Manager",
                                    note,
                                    now(),
                                    oos,
                                    status,
                                    None if sent else now(),
                                    now(),
                                    r.id
                                )
                            )
                            c.commit()
                            c.close()

                            eq2 = equipment()
                            mask = eq2.serial_number.astype(str).str.upper() == str(r.serial_number).upper()
                            eq2.loc[mask, "out_of_service"] = oos
                            save_equipment(eq2)

                            audit(
                                r.case_number,
                                "WH Manager",
                                manager_name.strip(),
                                "Sent to Maintenance" if sent else "Fixed by Manager",
                                note
                            )
                            st.rerun()

        unlisted = query(
            """SELECT ar.case_number,ar.serial_number,ar.found_location,ar.brand,ar.model,ar.unit_number,
                      ar.status,i.operator_name,i.issue_found
               FROM asset_review ar
               LEFT JOIN inspections i ON i.case_number=ar.case_number
               WHERE ar.found_location=?
               ORDER BY ar.id DESC""",
            (u["location"],)
        )
        if not unlisted.empty:
            st.markdown("### New Machine Requests")
            st.dataframe(humanize_columns(unlisted), hide_index=True, use_container_width=True)


    with tab_reports:
        daily_report_tab(u, "manager", all_locations=False)

    with tab_history:
        case_machine_history_tab(u, "manager", all_locations=False)


def corrective_actions(u):
    hour = datetime.now().hour
    greeting = "Good Morning" if hour < 12 else ("Good Afternoon" if hour < 17 else "Good Evening")
    st.header(f"{greeting}, {u['full_name']} 👋")
    st.caption("Maintenance")

    tab_work, tab_reports, tab_history = st.tabs(
        ["🔧 My Daily Work", "📋 Reports", "🕘 Case & Machine History"]
    )

    # =========================================================
    # TAB 1 — MY DAILY WORK
    # =========================================================
    with tab_work:
        repair_cases = query(
            """SELECT * FROM inspections
               WHERE status IN ('Repair Not Started','Repairing')
               ORDER BY id DESC"""
        )

        all_locations = sorted(
            [x for x in equipment().location.dropna().astype(str).unique() if x]
        )

        work_location = st.selectbox(
            "Location",
            all_locations,
            index=0 if all_locations else None,
            key="maintenance_work_location"
        ) if all_locations else ""

        scoped = repair_cases[
            repair_cases.location == work_location
        ].copy() if work_location else repair_cases.copy()

        waiting = scoped[scoped.status == "Repair Not Started"].copy()
        repairing = scoped[scoped.status == "Repairing"].copy()

        if work_location:
            today_loc = query(
                """SELECT * FROM inspections
                   WHERE inspection_date=? AND location=?
                   ORDER BY id DESC""",
                (str(date.today()), work_location)
            )
        else:
            today_loc = query(
                """SELECT * FROM inspections
                   WHERE inspection_date=?
                   ORDER BY id DESC""",
                (str(date.today()),)
            )

        fixed_today = today_loc[
            today_loc.status.fillna("").astype(str).str.startswith("Fixed by")
        ]
        delayed_today = today_loc[
            today_loc.repair_delay_note.fillna("").astype(str).str.strip() != ""
        ]

        st.markdown(f"### {work_location or 'All Locations'} — Today's Maintenance")
        a,b,c,d = st.columns(4)
        a.metric("Repair Not Started", len(waiting))
        b.metric("Repairing", len(repairing))
        c.metric("Fixed Today", len(fixed_today))
        d.metric("Need More Time", len(delayed_today))

        st.markdown("### Waiting for Repair")
        if waiting.empty:
            st.success("No repairs are waiting at this location.")
        else:
            for _, r in waiting.iterrows():
                with st.expander(
                    f"{r.case_number} • Serial {r.serial_number} • "
                    f"Shift {r.get('shift','—') or '—'} • Driver {r.operator_name}",
                    expanded=False
                ):
                    issues = query(
                        "SELECT component,issue_type,driver_note FROM issues WHERE case_number=?",
                        (r.case_number,)
                    )
                    st.dataframe(
                        humanize_columns(issues),
                        hide_index=True,
                        use_container_width=True
                    )
                    st.write(f"**Manager:** {r.manager_name or '—'}")
                    st.write(f"**Manager Note:** {r.manager_note or '—'}")

                    if st.button(
                        "▶ START REPAIR",
                        key=f"start_{r.id}",
                        type="primary"
                    ):
                        c = conn()
                        c.execute(
                            """UPDATE inspections
                               SET maintenance_name=?,repair_started_at=?,
                                   status='Repairing',updated_at=?
                               WHERE id=?""",
                            (u["full_name"], now(), now(), r.id)
                        )
                        c.commit()
                        c.close()

                        audit(
                            r.case_number,
                            "Maintenance",
                            u["full_name"],
                            "Repair Started",
                            ""
                        )
                        st.rerun()

        st.markdown("### Repairing")
        if repairing.empty:
            st.info("No repairs are currently in progress at this location.")
        else:
            for _, r in repairing.iterrows():
                with st.expander(
                    f"{r.case_number} • Serial {r.serial_number} • "
                    f"Shift {r.get('shift','—') or '—'} • Driver {r.operator_name}",
                    expanded=False
                ):
                    issues = query(
                        "SELECT component,issue_type,driver_note FROM issues WHERE case_number=?",
                        (r.case_number,)
                    )
                    st.dataframe(
                        humanize_columns(issues),
                        hide_index=True,
                        use_container_width=True
                    )

                    note = st.text_area(
                        "Maintenance Note / What was done? *",
                        key=f"repair_note_{r.id}"
                    )

                    photo = st.file_uploader(
                        "After-repair photo (optional)",
                        type=["jpg","jpeg","png"],
                        key=f"repair_photo_{r.id}"
                    )

                    if st.button(
                        "✅ FIXED",
                        key=f"fixed_{r.id}",
                        type="primary"
                    ):
                        if not note.strip():
                            st.error("Please enter the maintenance note.")
                        else:
                            pp = save_photo(photo, f"{r.case_number}_REPAIR")
                            c = conn()
                            c.execute(
                                """UPDATE inspections
                                   SET corrective_action=?,repair_completed_at=?,repair_photo=?,
                                       status=?,closed_at=?,updated_at=?
                                   WHERE id=?""",
                                (
                                    note,
                                    now(),
                                    pp,
                                    f"Fixed by {u['full_name']}",
                                    now(),
                                    now(),
                                    r.id
                                )
                            )
                            c.commit()
                            c.close()

                            audit(
                                r.case_number,
                                "Maintenance",
                                u["full_name"],
                                "Fixed by Maintenance",
                                note
                            )
                            st.rerun()

                    st.markdown("**Need more time?**")
                    delay_reason = st.selectbox(
                        "Reason",
                        DELAY_REASONS,
                        key=f"delay_reason_{r.id}"
                    )
                    delay_note = st.text_area(
                        "Delay Note *",
                        key=f"delay_note_{r.id}"
                    )

                    if st.button(
                        "⏳ NEED MORE TIME",
                        key=f"delay_{r.id}"
                    ):
                        if not delay_note.strip():
                            st.error("Please enter a short delay note.")
                        else:
                            c = conn()
                            c.execute(
                                """UPDATE inspections
                                   SET repair_delay_reason=?,repair_delay_note=?,updated_at=?
                                   WHERE id=?""",
                                (delay_reason, delay_note, now(), r.id)
                            )
                            c.commit()
                            c.close()

                            audit(
                                r.case_number,
                                "Maintenance",
                                u["full_name"],
                                "Repair Needs More Time",
                                f"{delay_reason}; {delay_note}"
                            )
                            st.success("Update saved. Case remains Repairing.")

    # =========================================================
    # TAB 2 — REPORTS
    # =========================================================
    with tab_reports:
        st.subheader("📋 Daily Logs & Reports")

        date_rows = query(
            """SELECT DISTINCT inspection_date
               FROM inspections
               ORDER BY inspection_date DESC"""
        )
        if date_rows.empty:
            st.info("No inspection records.")
        else:
            available_dates = [str(d) for d in date_rows.inspection_date.dropna().tolist()]

            f1,f2,f3 = st.columns(3)

            selected_date = f1.selectbox(
                "Date",
                available_dates,
                key="maintenance_report_date"
            )

            locations = ["All"] + sorted(
                [x for x in equipment().location.dropna().astype(str).unique() if x]
            )
            selected_location = f2.selectbox(
                "Location",
                locations,
                key="maintenance_report_location"
            )

            selected_shift = f3.selectbox(
                "Shift",
                ["All Shifts","Morning","Afternoon","Overnight"],
                key="maintenance_report_shift"
            )

            report_sql = "SELECT * FROM inspections WHERE inspection_date=?"
            report_params = [str(selected_date)]
            if selected_location != "All":
                report_sql += " AND location=?"
                report_params.append(selected_location)
            if selected_shift != "All Shifts":
                report_sql += " AND shift=?"
                report_params.append(selected_shift)
            day = query(report_sql + " ORDER BY id DESC", tuple(report_params))

            inspections_completed = len(
                day.drop_duplicates(subset=["serial_number","shift"])
            ) if len(day) else 0

            issues_reported = int(
                day.issue_found.fillna(0).astype(int).sum()
            ) if len(day) else 0

            sent_to_maintenance = len(
                day[
                    day.manager_decision.fillna("").astype(str) == "Repair Required"
                ]
            )

            fixed_by_maintenance = len(
                day[
                    day.status.fillna("").astype(str).str.startswith("Fixed by") &
                    day.maintenance_name.fillna("").astype(str).str.strip().ne("")
                ]
            )

            a,b,c,d = st.columns(4)
            a.metric("Inspections Completed", inspections_completed)
            b.metric("Issues Reported", issues_reported)
            c.metric("Sent to Maintenance", sent_to_maintenance)
            d.metric("Fixed by Maintenance", fixed_by_maintenance)

            st.markdown("### Shift Summary")

            shift_rows = []
            summary_sql = "SELECT * FROM inspections WHERE inspection_date=?"
            summary_params = [str(selected_date)]
            if selected_location != "All":
                summary_sql += " AND location=?"
                summary_params.append(selected_location)
            base_day = query(summary_sql + " ORDER BY id DESC", tuple(summary_params))

            for sh in ["Morning","Afternoon","Overnight"]:
                sh_day = base_day[
                    base_day["shift"].fillna("") == sh
                ].copy()

                shift_rows.append({
                    "Shift": sh,
                    "Inspections Completed": len(
                        sh_day.drop_duplicates(subset=["serial_number","shift"])
                    ) if len(sh_day) else 0,
                    "Issues Reported": int(
                        sh_day.issue_found.fillna(0).astype(int).sum()
                    ) if len(sh_day) else 0,
                    "Sent to Maintenance": len(
                        sh_day[
                            sh_day.manager_decision.fillna("").astype(str) == "Repair Required"
                        ]
                    ),
                    "Fixed by Maintenance": len(
                        sh_day[
                            sh_day.status.fillna("").astype(str).str.startswith("Fixed by") &
                            sh_day.maintenance_name.fillna("").astype(str).str.strip().ne("")
                        ]
                    )
                })

            st.dataframe(
                pd.DataFrame(shift_rows),
                hide_index=True,
                use_container_width=True
            )

            log_cols = [
                "case_number","inspection_date","shift","location","serial_number",
                "unit_number","operator_name","issue_found","status","created_at"
            ]
            log = day[[c for c in log_cols if c in day.columns]].copy()

            if "issue_found" in log.columns:
                log["issue_found"] = (
                    log["issue_found"].fillna(0).astype(int).map({0:"No",1:"Yes"})
                )
                log = log.rename(columns={"issue_found":"issue_reported"})

            log = log.rename(columns={
                "operator_name":"driver",
                "created_at":"inspection_time"
            })

            excel_data = build_excel_bytes([
                ("Daily Log", humanize_columns(log)),
                ("Shift Summary", pd.DataFrame(shift_rows)),
            ])

            st.download_button(
                "⬇️ Download Daily Log",
                excel_data,
                f"maintenance_daily_log_{selected_date}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="maintenance_daily_download"
            )

    # =========================================================
    # TAB 3 — CASE & MACHINE HISTORY
    # =========================================================
    with tab_history:
        st.subheader("🕘 View Case History")

        issue_cases = query(
            """SELECT * FROM inspections
               WHERE issue_found=1
               ORDER BY id DESC"""
        )

        if issue_cases.empty:
            st.info("No issue cases are available.")
        else:
            h1,h2,h3,h4 = st.columns(4)

            history_location = h1.selectbox(
                "Location",
                ["All"] + sorted(
                    [x for x in issue_cases.location.dropna().astype(str).unique() if x]
                ),
                key="maintenance_history_location"
            )

            history_shift = h2.selectbox(
                "Shift",
                ["All Shifts","Morning","Afternoon","Overnight"],
                key="maintenance_history_shift"
            )

            history_status = h3.selectbox(
                "Status",
                ["All"] + sorted(
                    [x for x in issue_cases.status.dropna().astype(str).unique() if x]
                ),
                key="maintenance_history_status"
            )

            history_oos = h4.selectbox(
                "Out of Service",
                ["All","Yes","No"],
                key="maintenance_history_oos"
            )

            hist = issue_cases.copy()

            if history_location != "All":
                hist = hist[hist.location == history_location]

            if history_shift != "All Shifts":
                hist = hist[hist["shift"].fillna("") == history_shift]

            if history_status != "All":
                hist = hist[hist.status == history_status]

            if history_oos != "All":
                hist = hist[hist.out_of_service.fillna("No") == history_oos]

            search = st.text_input(
                "Search Case # / Serial # / Unit # / Driver",
                key="maintenance_history_search"
            ).strip().lower()

            if search:
                hist = hist[
                    hist.case_number.astype(str).str.lower().str.contains(search, na=False) |
                    hist.serial_number.astype(str).str.lower().str.contains(search, na=False) |
                    hist.unit_number.astype(str).str.lower().str.contains(search, na=False) |
                    hist.operator_name.astype(str).str.lower().str.contains(search, na=False)
                ]

            history_cols = [
                "case_number","inspection_date","shift","location",
                "serial_number","unit_number","operator_name",
                "status","out_of_service"
            ]

            st.dataframe(
                humanize_columns(
                    hist[[c for c in history_cols if c in hist.columns]]
                ),
                hide_index=True,
                use_container_width=True
            )

            if not hist.empty:
                selected_case = st.selectbox(
                    "Select Issue Case",
                    hist.case_number.tolist(),
                    key="maintenance_history_case"
                )

                rec = hist[
                    hist.case_number == selected_case
                ].iloc[0]

                st.markdown(
                    f"**Case {selected_case}** • "
                    f"Shift **{rec.get('shift','—') or '—'}** • "
                    f"Serial **{rec.serial_number}** • "
                    f"Driver **{rec.operator_name}** • "
                    f"Status **{rec.status}**"
                )

                ii = query(
                    "SELECT component,issue_type,driver_note,created_at FROM issues WHERE case_number=?",
                    (selected_case,)
                )

                if not ii.empty:
                    st.markdown("**Reported Issue(s)**")
                    st.dataframe(
                        humanize_columns(ii),
                        hide_index=True,
                        use_container_width=True
                    )

                timeline = query(
                    """SELECT event_time,actor_role,actor_name,event_type,details
                       FROM audit
                       WHERE case_number=?
                       ORDER BY event_time""",
                    (selected_case,)
                )

                if not timeline.empty:
                    reported_issues = query(
                        "SELECT driver_note FROM issues WHERE case_number=?",
                        (selected_case,)
                    )

                    readable = []
                    for _, tr in timeline.iterrows():
                        event = str(tr.event_type or "")
                        raw = str(tr.details or "").strip()

                        if event == "Inspection Submitted":
                            notes = [
                                str(x).strip()
                                for x in reported_issues.driver_note.tolist()
                                if str(x or "").strip()
                            ]
                            detail = "; ".join(notes) if notes else "—"

                        elif event in ("Need More Time","Repair Needs More Time"):
                            detail = raw.split("; ",1)[1] if "; " in raw else (raw or "—")

                        else:
                            detail = raw if raw else "—"

                        readable.append(detail)

                    timeline = timeline.copy()
                    timeline["details"] = readable

                    st.markdown("**Case Timeline — Full Lifecycle**")
                    st.dataframe(
                        humanize_columns(timeline),
                        hide_index=True,
                        use_container_width=True
                    )

        st.divider()
        st.subheader("🚜 Machine History")

        eq = equipment()
        active_eq = eq[
            ~eq.status.astype(str).str.lower().isin(["inactive","retired","deleted"])
        ].copy()

        mh_location = st.selectbox(
            "Machine Location",
            ["All"] + sorted(
                [x for x in active_eq.location.dropna().astype(str).unique() if x]
            ),
            key="maintenance_machine_history_location"
        )

        if mh_location != "All":
            active_eq = active_eq[
                active_eq.location == mh_location
            ].copy()

        machine_labels = []
        machine_lookup = {}

        for _, er in active_eq.iterrows():
            label = (
                f"{er.serial_number} • Unit {er.unit_number or '—'} • "
                f"{er.brand} • {er.location}"
            )
            machine_labels.append(label)
            machine_lookup[label] = er

        selected_machine = st.selectbox(
            "Select Machine",
            machine_labels,
            index=None,
            placeholder="Choose a machine",
            key="maintenance_machine_history_select"
        )

        if selected_machine:
            er = machine_lookup[selected_machine]
            serial = str(er.serial_number)

            mh_all = query(
                """SELECT * FROM inspections
                   WHERE UPPER(serial_number)=UPPER(?)
                   ORDER BY inspection_date DESC, created_at DESC, id DESC""",
                (serial,)
            )

            mh_issues = mh_all[
                mh_all.issue_found.fillna(0).astype(int) == 1
            ].copy()

            fixed_mask = (
                mh_issues.status.fillna("").astype(str).str.startswith("Fixed by")
            )

            m1,m2,m3,m4 = st.columns(4)
            m1.metric("Total Inspections", len(mh_all))
            m2.metric("Total Tickets", len(mh_issues))
            m3.metric("Fixed Tickets", int(fixed_mask.sum()))
            m4.metric("Open Tickets", int((~fixed_mask).sum()))

            st.markdown(
                f"**Serial:** {er.serial_number}  •  "
                f"**Unit:** {er.unit_number or '—'}  •  "
                f"**Brand:** {er.brand or '—'}  •  "
                f"**Model:** {er.model or '—'}  •  "
                f"**Location:** {er.location}"
            )

            st.markdown("#### Inspection History")

            if mh_all.empty:
                st.info("No inspections have been recorded for this machine.")
            else:
                inspect_cols = [
                    "inspection_date","shift","case_number",
                    "operator_name","issue_found","status"
                ]

                mh_inspect = mh_all[
                    [c for c in inspect_cols if c in mh_all.columns]
                ].copy()

                if "issue_found" in mh_inspect.columns:
                    mh_inspect["issue_found"] = (
                        mh_inspect["issue_found"].fillna(0).astype(int)
                        .map({0:"No",1:"Yes"})
                    )
                    mh_inspect = mh_inspect.rename(
                        columns={"issue_found":"issue"}
                    )

                st.dataframe(
                    humanize_columns(mh_inspect),
                    hide_index=True,
                    use_container_width=True
                )

            st.markdown("#### Ticket History")

            if mh_issues.empty:
                st.success("No issue tickets have been recorded for this machine.")
            else:
                rows = []

                for _, mr in mh_issues.iterrows():
                    mi = query(
                        "SELECT component,issue_type FROM issues WHERE case_number=?",
                        (mr.case_number,)
                    )

                    issue_text = "; ".join(
                        [f"{x.component} — {x.issue_type}" for _, x in mi.iterrows()]
                    )

                    rows.append({
                        "inspection_date": mr.inspection_date,
                        "shift": mr.get("shift",""),
                        "case_number": mr.case_number,
                        "operator_name": mr.operator_name,
                        "issue": issue_text,
                        "status": mr.status
                    })

                st.dataframe(
                    humanize_columns(pd.DataFrame(rows)),
                    hide_index=True,
                    use_container_width=True
                )

def compliance_actions(u):
    hour = datetime.now().hour
    greeting = "Good Morning" if hour < 12 else ("Good Afternoon" if hour < 17 else "Good Evening")
    st.header(f"{greeting}, {u['full_name']} 👋")
    st.caption("Compliance • All Locations")

    tab_new, tab_reports, tab_history = st.tabs(
        ["➕ New Machine Requests", "📋 Reports", "🕘 Case & Machine History"]
    )

    with tab_new:
        q = query("""SELECT ar.*,i.operator_name,i.issue_found,i.status AS inspection_status
                     FROM asset_review ar
                     LEFT JOIN inspections i ON i.case_number=ar.case_number
                     WHERE ar.status IN ('Waiting Compliance Review','Waiting IT Location Update')
                     ORDER BY ar.id DESC""")

        eq = equipment()

        if q.empty:
            st.success("No new-machine requests are waiting.")
        else:
            for _, r in q.iterrows():
                title = (
                    f"{r.case_number} • Serial {r.serial_number} • "
                    f"Reported at {r.found_location} • Driver {r.operator_name}"
                )

                with st.expander(title, expanded=False):
                    matches = eq[
                        eq.serial_number.str.upper() == str(r.serial_number).upper()
                    ]

                    if r.status == "Waiting IT Location Update":
                        st.warning(
                            f"Location update requested: "
                            f"{r.registered_location} → {r.requested_location}"
                        )
                        st.caption("Waiting for IT to update the asset location.")
                        continue

                    if matches.empty:
                        st.success(
                            "Serial is not in Asset Master. "
                            "This can be added as a new machine."
                        )

                        locations = sorted([x for x in eq.location.unique() if x])

                        target = st.selectbox(
                            "Add to Location",
                            locations,
                            index=locations.index(r.found_location)
                            if r.found_location in locations else 0,
                            key=f"add_loc_{r.id}"
                        )

                        c1, c2, c3 = st.columns(3)

                        brand = c1.text_input(
                            "Brand",
                            value=r.brand or "",
                            key=f"brand_{r.id}"
                        )

                        model = c2.text_input(
                            "Model",
                            value=r.model or "",
                            key=f"model_{r.id}"
                        )

                        unit = c3.text_input(
                            "Unit #",
                            value=r.unit_number or "",
                            key=f"unit_{r.id}"
                        )

                        if st.button(
                            "➕ ADD MACHINE",
                            key=f"add_machine_{r.id}",
                            type="primary"
                        ):
                            new = pd.DataFrame([{
                                "serial_number": r.serial_number,
                                "unit_number": unit,
                                "location": target,
                                "brand": brand,
                                "description": "",
                                "model": model,
                                "status": "Active",
                                "building": "",
                                "out_of_service": "No"
                            }])

                            save_equipment(
                                pd.concat([eq, new], ignore_index=True)
                            )

                            c = conn()

                            c.execute("""UPDATE asset_review
                                         SET compliance_decision='New Machine Added',
                                             compliance_note=?,
                                             brand=?,model=?,unit_number=?,
                                             requested_location=?,
                                             status='Completed',
                                             updated_at=?
                                         WHERE id=?""",
                                      (
                                          f"Added to {target}",
                                          brand,
                                          model,
                                          unit,
                                          target,
                                          now(),
                                          r.id
                                      ))

                            c.execute("""INSERT INTO asset_movements(
                                         serial_number,from_location,to_location,
                                         case_number,approved_by,changed_at)
                                         VALUES(?,?,?,?,?,?)""",
                                      (
                                          r.serial_number,
                                          "",
                                          target,
                                          r.case_number,
                                          u["full_name"],
                                          now()
                                      ))

                            c.commit()
                            c.close()

                            audit(
                                r.case_number,
                                "Compliance",
                                u["full_name"],
                                "New Machine Added",
                                f"Serial {r.serial_number} -> {target}"
                            )

                            st.rerun()

                    else:
                        current = matches.iloc[0].location

                        if current == r.found_location:
                            st.warning(
                                f"Serial already exists at **{current}**. "
                                "A duplicate cannot be added."
                            )

                            if st.button(
                                "MARK REVIEWED",
                                key=f"reviewed_{r.id}"
                            ):
                                c = conn()

                                c.execute("""UPDATE asset_review
                                             SET registered_location=?,
                                                 compliance_decision='Already Exists',
                                                 status='Completed',
                                                 updated_at=?
                                             WHERE id=?""",
                                          (current, now(), r.id))

                                c.commit()
                                c.close()

                                audit(
                                    r.case_number,
                                    "Compliance",
                                    u["full_name"],
                                    "New Machine Request Reviewed",
                                    f"Serial already exists at {current}"
                                )

                                st.rerun()

                        else:
                            st.warning(
                                f"Serial is registered at **{current}**, "
                                f"but was reported at **{r.found_location}**."
                            )

                            st.info(
                                "Do not create a duplicate. "
                                "Send a location update request to IT."
                            )

                            locations = sorted(
                                [x for x in eq.location.unique() if x]
                            )

                            target = st.selectbox(
                                "Requested Location",
                                locations,
                                index=locations.index(r.found_location)
                                if r.found_location in locations else 0,
                                key=f"req_loc_{r.id}"
                            )

                            note = st.text_area(
                                "Compliance note",
                                key=f"req_note_{r.id}"
                            )

                            if st.button(
                                "SEND LOCATION UPDATE TO IT",
                                key=f"send_it_{r.id}",
                                type="primary"
                            ):
                                c = conn()

                                c.execute("""UPDATE asset_review
                                             SET registered_location=?,
                                                 requested_location=?,
                                                 compliance_decision='Location Update Requested',
                                                 compliance_note=?,
                                                 status='Waiting IT Location Update',
                                                 updated_at=?
                                             WHERE id=?""",
                                          (
                                              current,
                                              target,
                                              note,
                                              now(),
                                              r.id
                                          ))

                                c.commit()
                                c.close()

                                audit(
                                    r.case_number,
                                    "Compliance",
                                    u["full_name"],
                                    "Location Update Requested",
                                    f"{current} -> {target}; {note}"
                                )

                                st.rerun()


    with tab_reports:
        daily_report_tab(u, "compliance", all_locations=True)

    with tab_history:
        case_machine_history_tab(u, "compliance", all_locations=True)



def it_actions(u):
    hour = datetime.now().hour
    greeting = "Good Morning" if hour < 12 else ("Good Afternoon" if hour < 17 else "Good Evening")
    st.header(f"{greeting}, {u['full_name']} 👋")
    st.caption("IT Administration")

    tab_staff, tab_assets, tab_requests, tab_overview = st.tabs(
        ["👥 Staff Management", "🚜 Asset Management",
         "📍 Machine Location Update", "📊 System Overview"]
    )

    with tab_staff:
        st.subheader("Staff Management")
        st.caption("WH Managers are assigned to one warehouse. Maintenance, Compliance and IT have All Locations access.")

        users = query("SELECT * FROM users ORDER BY role,full_name")
        if users.empty:
            st.info("No staff users found.")
        else:
            show = users[["full_name","username","role","location","active"]].copy()
            show["location"] = show.apply(
                lambda r: r["location"] if r["role"]=="WH Manager" else "All Locations", axis=1
            )
            show["active"] = show["active"].map({1:"Active",0:"Inactive"}).fillna("Active")
            st.dataframe(humanize_columns(show), hide_index=True, use_container_width=True)

        st.markdown("### Add Staff")

        # Role is intentionally outside the form so Streamlit reruns immediately
        # when IT changes the role and can show the correct location control.
        role=st.selectbox(
            "Role *",
            ["WH Manager","Maintenance","Compliance","IT"],
            key="it_add_staff_role_v26"
        )
        locations=sorted([x for x in equipment().location.dropna().unique() if x])

        if role=="WH Manager":
            loc=st.selectbox(
                "Assigned Warehouse *",
                locations,
                key="it_add_staff_location_v26"
            )
            st.caption("Access: This warehouse only")
        else:
            loc="ALL"
            st.text_input(
                "Location Access",
                value="All Locations",
                disabled=True,
                key="it_add_staff_all_locations_v26"
            )

        with st.form("it_add_staff_v26", clear_on_submit=True):
            a,b=st.columns(2)
            full=a.text_input("Name *")
            username=b.text_input("Username *")
            password=st.text_input("Temporary Password *",type="password")
            add=st.form_submit_button("ADD STAFF",type="primary")

        if add:
            if not full.strip() or not username.strip() or not password:
                st.error("Name, Username and Temporary Password are required.")
            elif role == "WH Manager" and not loc:
                st.error("Assigned Warehouse is required for a WH Manager.")
            else:
                # Never trust a UI value for global roles.
                if role in ("Maintenance", "Compliance", "IT"):
                    loc = "ALL"
                salt=secrets.token_hex(16)
                h=hashlib.pbkdf2_hmac("sha256",password.encode(),salt.encode(),220000).hex()
                try:
                    c=conn()
                    c.execute("""INSERT INTO users(
                                 username,password_salt,password_hash,full_name,role,location,email,
                                 active,activated,must_change_password,temporary_password
                                 )
                                 VALUES(?,?,?,?,?,?,?,1,1,1,TRUE)""",
                              (username.strip(),salt,h,full.strip(),role,loc,""))
                    c.commit();c.close()
                    audit("","IT",u["full_name"],"Staff User Added",
                          f"Name={full.strip()}; Username={username.strip()}; Role={role}; Location={loc}")
                    st.success(f"{full.strip()} has been added.")
                    st.rerun()
                except Exception as ex:
                    st.error(str(ex))

        st.markdown("### Delete Staff")
        deletable=users[users.username.astype(str).str.lower()!="itadmin"].copy() if not users.empty else users
        if deletable.empty:
            st.info("No staff users available to delete.")
        else:
            labels={
                f"{r.full_name} • {r.username} • {r.role} • {r.location if r.role=='WH Manager' else 'All Locations'}":r
                for _,r in deletable.iterrows()
            }
            choice=st.selectbox("Staff User",list(labels.keys()),index=None,
                                placeholder="Choose a staff user",key="it_delete_staff")
            if choice:
                r=labels[choice]
                st.warning(f"This will permanently remove **{r.full_name} ({r.username})** from the system.")
                confirm=st.checkbox("I understand this user will be permanently deleted.",
                                    key=f"confirm_delete_user_{r.id}")
                if st.button("DELETE STAFF USER",type="primary",key=f"delete_user_{r.id}",
                             disabled=not confirm):
                    c=conn()
                    c.execute("DELETE FROM users WHERE id=?",(int(r.id),))
                    c.commit();c.close()
                    audit("","IT",u["full_name"],"Staff User Deleted",
                          f"Username={r.username}; Name={r.full_name}; Role={r.role}")
                    st.success("Staff user deleted.")
                    st.rerun()

    with tab_assets:
        st.subheader("Asset Management")
        eq=equipment()
        active=eq[~eq.status.astype(str).str.lower().isin(["deleted"])].copy()

        if active.empty:
            st.info("No active assets.")
        else:
            display_cols=[c for c in ["location","brand","model","serial_number","unit_number","status"] if c in active.columns]
            st.dataframe(humanize_columns(active[display_cols]),hide_index=True,use_container_width=True)

        st.markdown("### Delete Asset")
        labels={
            f"{r.serial_number} • Unit {r.unit_number or '—'} • {r.location} • {r.brand}":idx
            for idx,r in active.iterrows()
        }
        if labels:
            choice=st.selectbox("Asset",list(labels.keys()),index=None,
                                placeholder="Choose an asset",key="it_asset_delete")
            reason=st.text_area("Delete Reason *",key="it_asset_delete_reason")
            if st.button("DELETE ASSET",type="primary",key="it_delete_asset"):
                if not choice:
                    st.error("Please choose an asset.")
                elif not reason.strip():
                    st.error("Delete Reason is required.")
                else:
                    idx=labels[choice]
                    serial=eq.loc[idx,"serial_number"]
                    eq.loc[idx,"status"]="Deleted"
                    save_equipment(eq)
                    audit("","IT",u["full_name"],"Asset Deleted",
                          f"Serial={serial}; Reason={reason.strip()}")
                    st.success(f"Asset {serial} has been deleted.")
                    st.rerun()

    with tab_requests:
        st.subheader("Machine Location Update")
        st.caption("Update machines that have moved to a different warehouse after Compliance review.")

        req=query("""SELECT * FROM asset_review
                     WHERE status='Waiting IT Location Update'
                     ORDER BY id DESC""")
        eq=equipment()

        if req.empty:
            st.success("No location update requests are waiting.")
        else:
            for _,r in req.iterrows():
                with st.expander(
                    f"{r.case_number} • Serial {r.serial_number} • "
                    f"{r.registered_location} → {r.requested_location}"
                ):
                    st.write(f"**Current Location:** {r.registered_location}")
                    st.write(f"**Requested Location:** {r.requested_location}")
                    st.write(f"**Compliance Note:** {r.compliance_note or '—'}")

                    if st.button("APPROVE & UPDATE LOCATION",
                                 key=f"update_location_{r.id}",type="primary"):
                        mask=eq.serial_number.astype(str).str.upper()==str(r.serial_number).upper()
                        if not mask.any():
                            st.error("Serial was not found in the Asset Master.")
                        else:
                            old=eq.loc[mask,"location"].iloc[0]
                            eq.loc[mask,"location"]=r.requested_location
                            save_equipment(eq)

                            c=conn()
                            c.execute("""UPDATE asset_review
                                         SET status='Completed',updated_at=?
                                         WHERE id=?""",(now(),r.id))
                            c.execute("""INSERT INTO asset_movements(
                                         serial_number,from_location,to_location,
                                         case_number,approved_by,changed_at)
                                         VALUES(?,?,?,?,?,?)""",
                                      (r.serial_number,old,r.requested_location,
                                       r.case_number,u["full_name"],now()))
                            c.commit();c.close()

                            audit(r.case_number,"IT",u["full_name"],
                                  "Asset Location Updated",
                                  f"{old} -> {r.requested_location}")
                            st.success("Asset location updated.")
                            st.rerun()

    with tab_overview:
        st.subheader("System Overview")

        users=query("SELECT * FROM users")
        eq=equipment()
        active_assets=eq[~eq.status.astype(str).str.lower().isin(["deleted"])]
        pending=query("""SELECT * FROM asset_review
                         WHERE status='Waiting IT Location Update'""")

        a,b,c,d=st.columns(4)
        a.metric("Active Staff",len(users[users.active.fillna(1).astype(int)==1]) if not users.empty else 0)
        b.metric("Active Machines",len(active_assets))
        c.metric("Active Locations",active_assets.location.dropna().nunique())
        d.metric("Open IT Requests",len(pending))

        st.markdown("### Assets by Location")
        if active_assets.empty:
            st.info("No active assets.")
        else:
            by_loc=(active_assets.groupby("location",dropna=False)
                    .size().reset_index(name="Machines")
                    .rename(columns={"location":"Location"}))
            st.dataframe(by_loc,hide_index=True,use_container_width=True)

        st.markdown("### Staff by Role")
        if users.empty:
            st.info("No staff users.")
        else:
            by_role=(users.groupby("role",dropna=False)
                     .size().reset_index(name="Staff")
                     .rename(columns={"role":"Role"}))
            st.dataframe(by_role,hide_index=True,use_container_width=True)


def normalize_user_access(u):
    """Enforce role/location scope regardless of what is stored in the user record."""
    if not u:
        return u

    user = dict(u)
    role = str(user.get("role") or "").strip()

    if role in ("Maintenance", "Compliance", "Compliance & Audit", "IT"):
        user["location"] = "ALL"

    return user


def reminder_status():
    c = conn()
    due = c.execute("""
        SELECT COUNT(*) AS n
        FROM inspections
        WHERE status='Repairing'
          AND next_reminder_at IS NOT NULL
          AND next_reminder_at<=?
    """, (now(),)).fetchone()["n"]
    c.close()
    return due

init_db()

st.markdown("""
<div class='hero'>
    <h2>🏗️ 18W Forklift Safety Portal</h2>
    <div class='muted'>
        Daily inspection • repair control • compliance • role-based access
    </div>
</div>
""", unsafe_allow_html=True)

u = normalize_user_access(st.session_state.get("user"))

if u:
    # Keep the normalized access scope in session as well.
    st.session_state["user"] = u
    logout_button()

    if force_password_change(u):
        st.stop()

    role = u["role"]

    if role == "WH Manager":
        manager_actions(u)

    elif role == "Maintenance":
        corrective_actions(u)

    elif role in ("Compliance", "Compliance & Audit"):
        compliance_actions(u)

    elif role == "IT":
        it_actions(u)

    else:
        st.error("Invalid role.")

else:
    mode = st.sidebar.radio(
        "Open",
        ["🚜 Start Inspection", "🔐 Staff Login"]
    )

    if mode == "🚜 Start Inspection":
        run_driver()
    else:
        temporary_role_test_access()
        st.divider()
        staff_login()
    